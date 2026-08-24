import csv
import hashlib
import hmac
import ipaddress
import io
import json
import os
import secrets
import sqlite3
import threading
import time
from datetime import datetime, timedelta, timezone
from urllib.parse import urlsplit
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from flask import Flask, Response, jsonify, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get(
    "STATION_DB_PATH",
    os.path.join(BASE_DIR, "simple_station_swipes.db"),
)
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")
DASHBOARD_PATH = os.path.join(BASE_DIR, "dashboard.html")
BOOTSTRAP_ADMIN_CARD_ID = os.environ.get("STATION_BOOTSTRAP_ADMIN_CARD_ID", "").strip()
BOOTSTRAP_ADMIN_PASSWORD = os.environ.get("STATION_BOOTSTRAP_ADMIN_PASSWORD", "")
BOOTSTRAP_ADMIN_NAME = os.environ.get("STATION_BOOTSTRAP_ADMIN_NAME", "Bootstrap Admin")
BOOTSTRAP_ADMIN_USERNAME = os.environ.get(
    "STATION_BOOTSTRAP_ADMIN_USERNAME",
    "",
).strip().lower()
STATION_API_KEY = os.environ.get("STATION_API_KEY", "").strip()
SPACE_TIMEZONE_NAME = os.environ.get(
    "STATION_TIMEZONE",
    "America/Los_Angeles",
).strip()
try:
    SPACE_TIMEZONE = ZoneInfo(SPACE_TIMEZONE_NAME)
except ZoneInfoNotFoundError:
    SPACE_TIMEZONE = timezone.utc
    SPACE_TIMEZONE_NAME = "UTC"


def configured_hour(name, default, allow_24=False):
    try:
        value = int(os.environ.get(name, str(default)))
    except ValueError:
        return default
    maximum = 24 if allow_24 else 23
    return value if 0 <= value <= maximum else default


OPENING_HOUR = configured_hour("STATION_OPENING_HOUR", 12)
CLOSING_HOUR = configured_hour("STATION_CLOSING_HOUR", 17, allow_24=True)
PERSON_REF_SECRET = (
    os.environ.get("STATION_PERSON_REF_SECRET", "").strip()
    or STATION_API_KEY
    or secrets.token_urlsafe(32)
)
ROLE_LEVELS = {"volunteer": 1, "staff": 2, "admin": 3}
CERT_CONFIRM_SECONDS = 4
CERT_MODE_SECONDS = 40
SESSION_MAX_SECONDS = 12 * 60 * 60
SESSION_LIMIT_PER_ACCOUNT = 3
LOGIN_PAIR_WINDOW_SECONDS = 5 * 60
LOGIN_PAIR_ATTEMPT_LIMIT = 5
LOGIN_ACCOUNT_WINDOW_SECONDS = 15 * 60
LOGIN_ACCOUNT_ATTEMPT_LIMIT = 8
LOGIN_IP_WINDOW_SECONDS = 10 * 60
LOGIN_IP_ATTEMPT_LIMIT = 20
LOGIN_TRACKING_MAX_KEYS = 5000
SECURITY_AUDIT_COOLDOWN_SECONDS = 60
MAX_LOGIN_BODY_BYTES = 4096
PENDING_CARD_DAYS = 30
MAX_CSV_BYTES = 10 * 1024 * 1024
MAX_ID_LENGTH = 128
MAX_NAME_LENGTH = 200
MAX_EMAIL_LENGTH = 254
MAX_NOTES_LENGTH = 4000

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_CSV_BYTES
db_lock = threading.RLock()
session_lock = threading.RLock()
session_tokens = {}
login_attempts = {}
security_audit_cooldowns = {}
cert_modes = {}
DUMMY_PASSWORD_HASH = generate_password_hash(secrets.token_urlsafe(32))


def now_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def parse_iso(value):
    parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def elapsed_seconds(started_at, ended_at=None):
    try:
        start = (
            started_at
            if isinstance(started_at, datetime)
            else parse_iso(started_at)
        )
        end = ended_at or datetime.now(timezone.utc)
        return max(0, int((end - start).total_seconds()))
    except (TypeError, ValueError, OverflowError):
        return 0


def local_space_time(value=None):
    value = value or datetime.now(timezone.utc)
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(SPACE_TIMEZONE)


def space_is_open(value=None):
    local_time = local_space_time(value)
    decimal_hour = local_time.hour + (local_time.minute / 60)
    if OPENING_HOUR == CLOSING_HOUR:
        return True
    if OPENING_HOUR < CLOSING_HOUR:
        return OPENING_HOUR <= decimal_hour < CLOSING_HOUR
    return decimal_hour >= OPENING_HOUR or decimal_hour < CLOSING_HOUR


def display_hour(hour):
    if hour == 24:
        hour = 0
    suffix = "AM" if hour < 12 else "PM"
    display = hour % 12 or 12
    return f"{display}:00 {suffix}"


def opening_hours_label():
    return f"{display_hour(OPENING_HOUR)}-{display_hour(CLOSING_HOUR)}"


def after_hours_access_role(conn, card_id):
    row = conn.execute(
        """
        SELECT cards.designation, user_accounts.role, user_accounts.active
        FROM cards
        LEFT JOIN user_accounts ON user_accounts.card_id = cards.card_id
        WHERE cards.card_id = ?
        """,
        (card_id,),
    ).fetchone()
    if not row:
        return ""

    account_role = normalize_access_role(row["role"])
    if row["active"] and account_role in ("admin", "staff", "volunteer"):
        return account_role

    designation = normalize_access_role(row["designation"])
    return designation if designation in ("staff", "volunteer") else ""


def normalize_access_role(value):
    text = str(value or "").strip().lower()
    return text if text in ROLE_LEVELS else ""


def role_label(role):
    normalized = normalize_access_role(role)
    return normalized.title() if normalized else ""


def role_allows(actual_role, *allowed_roles):
    actual_level = ROLE_LEVELS.get(normalize_access_role(actual_role), 0)
    return any(
        actual_level >= ROLE_LEVELS.get(normalize_access_role(role), 999)
        for role in allowed_roles
    )


def token_from_request():
    return request.headers.get("X-Access-Token", "") or ""


def account_for_token(token):
    with session_lock:
        session = session_tokens.get(token)
        if session and session.get("expires_at", 0) <= time.time():
            session_tokens.pop(token, None)
            session = None

    if not session:
        return None

    with db_lock:
        conn = db_connect()
        row = conn.execute(
            """
            SELECT
                user_accounts.card_id,
                user_accounts.username,
                user_accounts.role,
                user_accounts.active,
                cards.name,
                cards.email
            FROM user_accounts
            JOIN cards ON cards.card_id = user_accounts.card_id
            WHERE user_accounts.card_id = ?
            """,
            (session["card_id"],),
        ).fetchone()
        conn.close()

    if not row or not row["active"]:
        with session_lock:
            session_tokens.pop(token, None)
        return None

    return dict(row)


def require_access(*allowed_roles):
    account = account_for_token(token_from_request())
    if account and role_allows(account["role"], *allowed_roles):
        return normalize_access_role(account["role"]), None

    if allowed_roles == ("admin",):
        message = "Admin login required"
    elif allowed_roles == ("staff",):
        message = "Staff or admin login required"
    else:
        message = "Authorized login required"

    return "", (jsonify({"ok": False, "error": message}), 403)


def require_station_api_key():
    provided_key = request.headers.get("X-Station-Key", "")
    if STATION_API_KEY and secrets.compare_digest(provided_key, STATION_API_KEY):
        return None

    return jsonify({"ok": False, "error": "Station API key required"}), 401


def invalidate_sessions_for_card(card_id, except_token=""):
    with session_lock:
        for token, session in list(session_tokens.items()):
            if session.get("card_id") == card_id and token != except_token:
                session_tokens.pop(token, None)


def client_ip_address():
    remote_address = request.remote_addr or "unknown"
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if remote_address in ("127.0.0.1", "::1") and forwarded_for:
        candidate = forwarded_for.split(",", 1)[0].strip()
        try:
            return str(ipaddress.ip_address(candidate))
        except ValueError:
            pass
    try:
        return str(ipaddress.ip_address(remote_address))
    except ValueError:
        return "unknown"


def login_attempt_scopes(login):
    client_ip = client_ip_address()
    normalized_login = str(login or "").strip().lower()
    return {
        "pair": ("pair", client_ip, normalized_login),
        "account": ("account", normalized_login),
        "ip": ("ip", client_ip),
    }


def login_limit_rules():
    return {
        "pair": (LOGIN_PAIR_ATTEMPT_LIMIT, LOGIN_PAIR_WINDOW_SECONDS),
        "account": (LOGIN_ACCOUNT_ATTEMPT_LIMIT, LOGIN_ACCOUNT_WINDOW_SECONDS),
        "ip": (LOGIN_IP_ATTEMPT_LIMIT, LOGIN_IP_WINDOW_SECONDS),
    }


def recent_login_attempts(key, window_seconds, current_time):
    cutoff = current_time - window_seconds
    attempts = [stamp for stamp in login_attempts.get(key, []) if stamp > cutoff]
    if attempts:
        login_attempts[key] = attempts
    else:
        login_attempts.pop(key, None)
    return attempts


def login_retry_after(login):
    current_time = time.time()
    longest_retry = 0
    scopes = login_attempt_scopes(login)
    with session_lock:
        for scope_name, (limit, window_seconds) in login_limit_rules().items():
            attempts = recent_login_attempts(
                scopes[scope_name],
                window_seconds,
                current_time,
            )
            if len(attempts) >= limit:
                retry_after = int(attempts[0] + window_seconds - current_time) + 1
                longest_retry = max(longest_retry, retry_after)
    return max(0, longest_retry)


def record_login_failure(login):
    current_time = time.time()
    scopes = login_attempt_scopes(login)
    with session_lock:
        for scope_name, (_, window_seconds) in login_limit_rules().items():
            key = scopes[scope_name]
            attempts = recent_login_attempts(key, window_seconds, current_time)
            attempts.append(current_time)
            login_attempts[key] = attempts

        if len(login_attempts) > LOGIN_TRACKING_MAX_KEYS:
            oldest_keys = sorted(
                login_attempts,
                key=lambda key: login_attempts[key][-1] if login_attempts[key] else 0,
            )
            for key in oldest_keys[: len(login_attempts) - LOGIN_TRACKING_MAX_KEYS]:
                login_attempts.pop(key, None)


def clear_login_failures(login):
    scopes = login_attempt_scopes(login)
    with session_lock:
        login_attempts.pop(scopes["pair"], None)
        login_attempts.pop(scopes["account"], None)


def should_log_security_event(action, login):
    client_ip = client_ip_address()
    normalized_login = str(login or "").strip().lower()
    keys = (
        (action, "ip", client_ip),
        (action, "account", normalized_login),
    )
    current_time = time.time()
    with session_lock:
        if any(
            current_time - security_audit_cooldowns.get(key, 0)
            < SECURITY_AUDIT_COOLDOWN_SECONDS
            for key in keys
        ):
            return False
        for key in keys:
            security_audit_cooldowns[key] = current_time
    return True


def prune_in_memory_auth_state():
    current_time = time.time()
    maximum_window = max(
        LOGIN_PAIR_WINDOW_SECONDS,
        LOGIN_ACCOUNT_WINDOW_SECONDS,
        LOGIN_IP_WINDOW_SECONDS,
    )
    with session_lock:
        for token, session in list(session_tokens.items()):
            if session.get("expires_at", 0) <= current_time:
                session_tokens.pop(token, None)
        for key, attempts in list(login_attempts.items()):
            recent_attempt_list = [
                stamp for stamp in attempts if stamp > current_time - maximum_window
            ]
            if recent_attempt_list:
                login_attempts[key] = recent_attempt_list
            else:
                login_attempts.pop(key, None)
        for key, logged_at in list(security_audit_cooldowns.items()):
            if current_time - logged_at >= SECURITY_AUDIT_COOLDOWN_SECONDS:
                security_audit_cooldowns.pop(key, None)


def validate_text(value, field, max_length, required=False):
    text = str(value or "").strip()
    if required and not text:
        raise ValueError(f"{field} is required")
    if len(text) > max_length:
        raise ValueError(f"{field} must be {max_length} characters or fewer")
    return text


def validate_email(value, required=False):
    email = validate_text(value, "email", MAX_EMAIL_LENGTH, required)
    if email and ("@" not in email or email.startswith("@") or email.endswith("@")):
        raise ValueError("email must be a valid email address")
    return email


def request_json():
    data = request.get_json(silent=True)
    return data if isinstance(data, dict) else {}


def request_host_is_allowed_for_origin(origin):
    try:
        origin_host = urlsplit(origin).netloc.lower()
    except ValueError:
        return False
    if not origin_host:
        return False

    allowed_hosts = {request.host.lower()}
    if (request.remote_addr or "") in ("127.0.0.1", "::1"):
        forwarded_host = request.headers.get("X-Forwarded-Host", "")
        if forwarded_host:
            allowed_hosts.add(forwarded_host.split(",", 1)[0].strip().lower())
    return origin_host in allowed_hosts


@app.before_request
def reject_cross_origin_writes():
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return None
    if not (request.path.startswith("/api/") or request.path == "/swipe"):
        return None
    origin = request.headers.get("Origin", "")
    if origin and not request_host_is_allowed_for_origin(origin):
        return jsonify({"ok": False, "error": "Cross-origin request rejected"}), 403
    return None


@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "no-referrer")
    response.headers.setdefault("Cross-Origin-Opener-Policy", "same-origin")
    response.headers.setdefault("Cross-Origin-Resource-Policy", "same-origin")
    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(), microphone=(), geolocation=()",
    )
    response.headers.setdefault("X-Permitted-Cross-Domain-Policies", "none")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; connect-src 'self'; img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; "
        "frame-ancestors 'none'; base-uri 'none'; form-action 'self'",
    )
    if (
        request.path.startswith("/api/")
        or request.path in ("/", "/dashboard")
        or request.path.endswith((".csv", ".xlsx"))
    ):
        response.headers["Cache-Control"] = "no-store"
    forwarded_https = (
        (request.remote_addr or "") in ("127.0.0.1", "::1")
        and request.headers.get("X-Forwarded-Proto", "").lower() == "https"
    )
    if request.is_secure or forwarded_https:
        response.headers.setdefault(
            "Strict-Transport-Security",
            "max-age=31536000; includeSubDomains",
        )
    return response


@app.errorhandler(RequestEntityTooLarge)
def upload_too_large(_error):
    return jsonify(
        {
            "ok": False,
            "error": f"Upload is too large; maximum size is {MAX_CSV_BYTES // (1024 * 1024)} MB",
        }
    ), 413


@app.errorhandler(Exception)
def unhandled_error(error):
    if isinstance(error, HTTPException):
        return error

    app.logger.exception("Unhandled request error")
    if request.path.startswith("/api/") or request.path == "/swipe":
        return jsonify({"ok": False, "error": "Internal server error"}), 500
    return Response("Internal server error", status=500, mimetype="text/plain")


def db_connect():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn


def ensure_column(conn, table, column, definition):
    columns = conn.execute(f"PRAGMA table_info({table})").fetchall()
    if column not in [row["name"] for row in columns]:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db():
    with db_lock:
        conn = db_connect()
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA synchronous = NORMAL")
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS stations (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                kind TEXT NOT NULL DEFAULT 'station',
                dashboard_visible INTEGER NOT NULL DEFAULT 1,
                requires_certification INTEGER NOT NULL DEFAULT 1,
                cert_override_active INTEGER NOT NULL DEFAULT 0,
                cert_override_by TEXT NOT NULL DEFAULT '',
                cert_override_updated_at TEXT NOT NULL DEFAULT '',
                cert_override_expires_at TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS active_people (
                card_id TEXT PRIMARY KEY,
                entered_at TEXT NOT NULL,
                entry_door_id TEXT NOT NULL,
                entry_door_name TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS active_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                card_id TEXT NOT NULL,
                station_id TEXT NOT NULL,
                station_name TEXT NOT NULL,
                started_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS people (
                bronco_id TEXT PRIMARY KEY,
                name TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL DEFAULT '',
                designation TEXT NOT NULL DEFAULT 'User',
                active INTEGER NOT NULL DEFAULT 1,
                notes TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS cards (
                card_id TEXT PRIMARY KEY,
                bronco_id TEXT NOT NULL DEFAULT '',
                name TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL DEFAULT '',
                designation TEXT NOT NULL DEFAULT 'User',
                active INTEGER NOT NULL DEFAULT 1,
                notes TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                FOREIGN KEY (bronco_id) REFERENCES people(bronco_id)
            );

            CREATE TABLE IF NOT EXISTS user_accounts (
                card_id TEXT PRIMARY KEY,
                username TEXT NOT NULL DEFAULT '',
                role TEXT NOT NULL DEFAULT 'Volunteer',
                password_hash TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (card_id) REFERENCES cards(card_id)
            );

            CREATE TABLE IF NOT EXISTS certifications (
                card_id TEXT NOT NULL,
                station_id TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                notes TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                granted_via TEXT NOT NULL DEFAULT 'dashboard',
                granted_by TEXT NOT NULL DEFAULT '',
                granted_at TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (card_id, station_id),
                FOREIGN KEY (card_id) REFERENCES cards(card_id),
                FOREIGN KEY (station_id) REFERENCES stations(id)
            );

            CREATE TABLE IF NOT EXISTS bronco_certifications (
                bronco_id TEXT NOT NULL,
                station_id TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                notes TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                PRIMARY KEY (bronco_id, station_id),
                FOREIGN KEY (bronco_id) REFERENCES people(bronco_id),
                FOREIGN KEY (station_id) REFERENCES stations(id)
            );

            CREATE TABLE IF NOT EXISTS certify_permissions (
                card_id TEXT NOT NULL,
                station_id TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (card_id, station_id),
                FOREIGN KEY (card_id) REFERENCES user_accounts(card_id),
                FOREIGN KEY (station_id) REFERENCES stations(id)
            );

            CREATE TABLE IF NOT EXISTS canvas_sync_tasks (
                bronco_id TEXT NOT NULL,
                card_id TEXT NOT NULL DEFAULT '',
                station_id TEXT NOT NULL,
                desired_active INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                created_by_card_id TEXT NOT NULL DEFAULT '',
                created_by_name TEXT NOT NULL DEFAULT '',
                completed_at TEXT NOT NULL DEFAULT '',
                completed_by_card_id TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (bronco_id, station_id),
                FOREIGN KEY (bronco_id) REFERENCES people(bronco_id),
                FOREIGN KEY (station_id) REFERENCES stations(id)
            );

            CREATE TABLE IF NOT EXISTS swipe_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                card_id TEXT NOT NULL,
                station_id TEXT NOT NULL,
                station_name TEXT NOT NULL,
                station_kind TEXT NOT NULL DEFAULT 'station',
                timestamp TEXT NOT NULL,
                action TEXT NOT NULL,
                allowed INTEGER NOT NULL DEFAULT 1,
                duration_seconds INTEGER,
                active_users INTEGER NOT NULL,
                warning TEXT NOT NULL DEFAULT '',
                details TEXT NOT NULL DEFAULT '',
                event_id TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS pending_card_dismissals (
                card_id TEXT PRIMARY KEY,
                ignored_through TEXT NOT NULL,
                ignored_through_event_id INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS warning_dismissals (
                swipe_event_id INTEGER PRIMARY KEY,
                dismissed_at TEXT NOT NULL,
                dismissed_by_card_id TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (swipe_event_id) REFERENCES swipe_events(id)
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                actor_card_id TEXT NOT NULL DEFAULT '',
                actor_name TEXT NOT NULL DEFAULT '',
                actor_role TEXT NOT NULL DEFAULT '',
                action TEXT NOT NULL,
                target_type TEXT NOT NULL DEFAULT '',
                target_id TEXT NOT NULL DEFAULT '',
                details TEXT NOT NULL DEFAULT ''
            );
            """
        )

        ensure_column(conn, "stations", "kind", "TEXT NOT NULL DEFAULT 'station'")
        ensure_column(
            conn,
            "stations",
            "dashboard_visible",
            "INTEGER NOT NULL DEFAULT 1",
        )
        ensure_column(conn, "swipe_events", "station_kind", "TEXT NOT NULL DEFAULT 'station'")
        ensure_column(conn, "swipe_events", "allowed", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "swipe_events", "warning", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "swipe_events", "details", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "swipe_events", "event_id", "TEXT NOT NULL DEFAULT ''")
        ensure_column(
            conn,
            "pending_card_dismissals",
            "ignored_through_event_id",
            "INTEGER NOT NULL DEFAULT 0",
        )
        conn.execute(
            """
            UPDATE swipe_events
            SET event_id = ''
            WHERE event_id != ''
              AND id NOT IN (
                  SELECT MIN(id)
                  FROM swipe_events
                  WHERE event_id != ''
                  GROUP BY event_id
              )
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_swipe_events_event_id
            ON swipe_events(event_id)
            WHERE event_id != ''
            """
        )
        ensure_column(conn, "cards", "bronco_id", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cards", "designation", "TEXT NOT NULL DEFAULT 'User'")
        ensure_column(conn, "user_accounts", "username", "TEXT NOT NULL DEFAULT ''")
        ensure_column(
            conn,
            "stations",
            "requires_certification",
            "INTEGER NOT NULL DEFAULT 1",
        )
        ensure_column(
            conn,
            "stations",
            "cert_override_active",
            "INTEGER NOT NULL DEFAULT 0",
        )
        ensure_column(conn, "stations", "cert_override_by", "TEXT NOT NULL DEFAULT ''")
        ensure_column(
            conn,
            "stations",
            "cert_override_updated_at",
            "TEXT NOT NULL DEFAULT ''",
        )
        ensure_column(
            conn,
            "stations",
            "cert_override_expires_at",
            "TEXT NOT NULL DEFAULT ''",
        )
        ensure_column(
            conn,
            "certifications",
            "granted_via",
            "TEXT NOT NULL DEFAULT 'dashboard'",
        )
        ensure_column(
            conn,
            "certifications",
            "granted_by",
            "TEXT NOT NULL DEFAULT ''",
        )
        ensure_column(
            conn,
            "certifications",
            "granted_at",
            "TEXT NOT NULL DEFAULT ''",
        )

        conn.execute("DROP INDEX IF EXISTS idx_certifications_card_station")
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_people_email
            ON people(email)
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_people_bronco_id_unique
            ON people(lower(bronco_id))
            WHERE bronco_id != ''
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_cards_bronco_id
            ON cards(bronco_id)
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_cards_bronco_id_unique
            ON cards(lower(bronco_id))
            WHERE bronco_id != ''
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_cards_email
            ON cards(email)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_user_accounts_role
            ON user_accounts(role)
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_user_accounts_username
            ON user_accounts(lower(username))
            WHERE username != ''
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_swipe_events_timestamp
            ON swipe_events(timestamp)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_swipe_events_pending_cards
            ON swipe_events(warning, card_id, id DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_warning_dismissals_event
            ON warning_dismissals(swipe_event_id)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_active_sessions_card
            ON active_sessions(card_id, station_id)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_active_sessions_station
            ON active_sessions(station_id)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_audit_log_timestamp
            ON audit_log(timestamp)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_canvas_sync_tasks_status
            ON canvas_sync_tasks(status, created_at)
            """
        )
        blank_usernames = conn.execute(
            """
            SELECT user_accounts.card_id, cards.email
            FROM user_accounts
            JOIN cards ON cards.card_id = user_accounts.card_id
            WHERE user_accounts.username = ''
            """
        ).fetchall()
        for row in blank_usernames:
            username = login_username_from_email(row["email"])
            if not username:
                continue
            duplicate = conn.execute(
                """
                SELECT card_id
                FROM user_accounts
                WHERE lower(username) = lower(?) AND card_id != ?
                """,
                (username, row["card_id"]),
            ).fetchone()
            if duplicate:
                username = row["card_id"]
            conn.execute(
                "UPDATE user_accounts SET username = ? WHERE card_id = ?",
                (username, row["card_id"]),
            )

        conn.execute("PRAGMA optimize")

        conn.commit()
        conn.close()


def seed_stations():
    if not os.path.exists(CONFIG_PATH):
        return

    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as file:
            config = json.load(file)
    except (OSError, json.JSONDecodeError) as error:
        raise RuntimeError(f"Unable to read station config: {error}") from error

    stations = config.get("stations")
    if not isinstance(stations, list):
        raise RuntimeError("config.json must contain a stations list")

    validated_stations = []
    seen_ids = set()
    for index, station in enumerate(stations, start=1):
        if not isinstance(station, dict):
            raise RuntimeError(f"Station {index} in config.json must be an object")
        try:
            station_id = validate_text(
                station.get("id"),
                f"station {index} id",
                MAX_ID_LENGTH,
                required=True,
            )
            station_name = validate_text(
                station.get("name"),
                f"station {index} name",
                MAX_NAME_LENGTH,
                required=True,
            )
        except ValueError as error:
            raise RuntimeError(str(error)) from error
        station_kind = str(station.get("kind", "station")).strip().lower()
        if station_kind not in ("door", "station"):
            raise RuntimeError(
                f"Station {station_id} has invalid kind {station_kind!r}"
            )
        normalized_id = station_id.lower()
        if normalized_id in seen_ids:
            raise RuntimeError(f"Duplicate station id in config.json: {station_id}")
        seen_ids.add(normalized_id)
        validated_stations.append((station_id, station_name, station_kind))

    with db_lock:
        conn = db_connect()
        for station_id, station_name, station_kind in validated_stations:
            conn.execute(
                """
                INSERT INTO stations (id, name, kind)
                VALUES (?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    name = excluded.name,
                    kind = excluded.kind
                """,
                (station_id, station_name, station_kind),
            )
        conn.commit()
        conn.close()


def bootstrap_admin():
    if not BOOTSTRAP_ADMIN_CARD_ID or not BOOTSTRAP_ADMIN_PASSWORD:
        return

    with db_lock:
        conn = db_connect()
        admin_count = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM user_accounts
            WHERE lower(role) = 'admin' AND active = 1
            """
        ).fetchone()["count"]

        if admin_count:
            if BOOTSTRAP_ADMIN_USERNAME:
                existing_bootstrap = conn.execute(
                    """
                    SELECT username
                    FROM user_accounts
                    WHERE card_id = ?
                    """,
                    (BOOTSTRAP_ADMIN_CARD_ID,),
                ).fetchone()
                if existing_bootstrap and (
                    not existing_bootstrap["username"]
                    or existing_bootstrap["username"] == BOOTSTRAP_ADMIN_CARD_ID
                ):
                    duplicate = conn.execute(
                        """
                        SELECT card_id
                        FROM user_accounts
                        WHERE lower(username) = lower(?) AND card_id != ?
                        """,
                        (BOOTSTRAP_ADMIN_USERNAME, BOOTSTRAP_ADMIN_CARD_ID),
                    ).fetchone()
                    if not duplicate:
                        conn.execute(
                            """
                            UPDATE user_accounts
                            SET username = ?, updated_at = ?
                            WHERE card_id = ?
                            """,
                            (
                                BOOTSTRAP_ADMIN_USERNAME,
                                now_iso(),
                                BOOTSTRAP_ADMIN_CARD_ID,
                            ),
                        )
                        conn.commit()
            conn.close()
            return

        timestamp = now_iso()
        conn.execute(
            """
            INSERT OR IGNORE INTO cards (
                card_id, name, email, designation, active, notes, updated_at
            )
            VALUES (?, ?, '', 'Staff', 1, 'Bootstrap admin account', ?)
            """,
            (BOOTSTRAP_ADMIN_CARD_ID, BOOTSTRAP_ADMIN_NAME, timestamp),
        )
        conn.execute(
            """
            INSERT INTO user_accounts (
                card_id, username, role, password_hash, active, updated_at
            )
            VALUES (?, ?, 'Admin', ?, 1, ?)
            ON CONFLICT(card_id) DO UPDATE SET
                username = excluded.username,
                role = 'Admin',
                password_hash = excluded.password_hash,
                active = 1,
                updated_at = excluded.updated_at
            """,
            (
                BOOTSTRAP_ADMIN_CARD_ID,
                BOOTSTRAP_ADMIN_USERNAME or BOOTSTRAP_ADMIN_CARD_ID,
                generate_password_hash(BOOTSTRAP_ADMIN_PASSWORD),
                timestamp,
            ),
        )
        add_audit_log(
            conn,
            {
                "card_id": BOOTSTRAP_ADMIN_CARD_ID,
                "name": BOOTSTRAP_ADMIN_NAME,
                "role": "Admin",
            },
            "bootstrap_admin_created",
            "account",
            BOOTSTRAP_ADMIN_CARD_ID,
            {"username": BOOTSTRAP_ADMIN_USERNAME or BOOTSTRAP_ADMIN_CARD_ID},
        )
        conn.commit()
        conn.close()


def infer_kind(station_id, station_name):
    text = f"{station_id} {station_name}".lower()
    return "door" if "door" in text else "station"


def override_is_effective(station):
    if not station or not station["cert_override_active"]:
        return False

    expires_at = station["cert_override_expires_at"]
    if not expires_at:
        return False

    try:
        return parse_iso(expires_at) > datetime.now(timezone.utc)
    except (TypeError, ValueError, OverflowError):
        return False


def station_override_warning(station, warning="", details=""):
    if not override_is_effective(station):
        return warning, details

    override_details = (
        "Certification override active until "
        f"{station['cert_override_expires_at']}; "
        f"enabled by {station['cert_override_by'] or 'unknown'}"
    )
    if warning:
        override_details += f"; original warning={warning}"
    if details:
        override_details += f"; {details}"

    return "cert_override_active", override_details


def get_station_info(conn, station_id, provided_name="", provided_kind=""):
    station = conn.execute(
        """
        SELECT
            id,
            name,
            kind,
            requires_certification,
            cert_override_active,
            cert_override_by,
            cert_override_updated_at,
            cert_override_expires_at
        FROM stations
        WHERE id = ?
        """,
        (station_id,),
    ).fetchone()

    if station:
        return {
            "station_id": station_id,
            "station_name": station["name"],
            "station_kind": station["kind"],
            "requires_certification": bool(station["requires_certification"]),
            "cert_override_active": override_is_effective(station),
            "cert_override_by": station["cert_override_by"] or "",
            "cert_override_updated_at": station["cert_override_updated_at"] or "",
            "cert_override_expires_at": station["cert_override_expires_at"] or "",
        }

    station_name = str(provided_name or station_id)[:MAX_NAME_LENGTH]
    normalized_kind = str(provided_kind or "").strip().lower()
    station_kind = (
        normalized_kind
        if normalized_kind in ("door", "station")
        else infer_kind(station_id, station_name)
    )

    conn.execute(
        """
        INSERT INTO stations (id, name, kind)
        VALUES (?, ?, ?)
        """,
        (station_id, station_name, station_kind),
    )

    return {
        "station_id": station_id,
        "station_name": station_name,
        "station_kind": station_kind,
        "requires_certification": True,
        "cert_override_active": False,
        "cert_override_by": "",
        "cert_override_updated_at": "",
        "cert_override_expires_at": "",
    }


def building_active_count(conn):
    row = conn.execute("SELECT COUNT(*) AS count FROM active_people").fetchone()
    return int(row["count"])


def parse_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    return str(value).strip().lower() in (
        "1",
        "1.0",
        "true",
        "yes",
        "y",
        "on",
        "active",
        "checked",
    )


def normalize_designation(value):
    text = str(value or "User").strip().lower()
    options = {
        "staff": "Staff",
        "volunteer": "Volunteer",
        "user": "User",
    }
    return options.get(text, "User")


def login_username_from_email(email):
    email = str(email or "").strip()
    if "@" not in email:
        return ""
    return email.split("@", 1)[0].strip().lower()


def csv_text_from_request():
    if request.files:
        file = next(iter(request.files.values()))
        try:
            return file.read().decode("utf-8-sig")
        except UnicodeDecodeError as error:
            raise ValueError("CSV must be UTF-8 encoded") from error

    data = request_json()
    return str(data.get("csv", data.get("text", "")))


def normalized_csv_rows(text):
    text = str(text or "")
    if not text.strip():
        return []

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample.splitlines()[0] else csv.excel

    rows = []
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    for raw in reader:
        row = {}
        for key, value in raw.items():
            if key is None:
                continue
            clean_key = str(key).strip().lower().replace(" ", "_").replace("-", "_")
            row[clean_key] = str(value or "").strip()
        rows.append(row)
    return rows


def row_value(row, *keys):
    for key in keys:
        clean_key = str(key).strip().lower().replace(" ", "_").replace("-", "_")
        value = row.get(clean_key, "")
        if value:
            return value
    return ""


CERTIFICATION_IMPORT_COLUMNS = (
    ("Sticker Making - In Person (1253702)", "1253702", "stickers"),
    ("Button Making - In Person (1253671)", "1253671", "buttons-stickers"),
    ("Leather Work In-person Certification (1253687)", "1253687", "leather-working"),
    ("3D Printing - In Person Component (1253668)", "1253668", "3d-printing"),
    ("Sewing Training - In Person (1253700)", "1253700", "sewing"),
    ("Soldering & Power Supply - In Person (1253701)", "1253701", "soldering"),
    ("Embroidery Training- In Person (1253680)", "1253680", "embroidery"),
    ("Vinyl Cutting - In Person (1253709)", "1253709", "vinyl"),
    ("Letterpress - In Person Component (1253691)", "1253691", "letter-press"),
)


def student_name_from_row(row):
    name = row_value(row, "student", "name", "person", "full_name")
    if "," not in name:
        return name

    last_name, first_name = (part.strip() for part in name.split(",", 1))
    return " ".join(part for part in (first_name, last_name) if part)


def certification_import_value(row, canvas_course_id):
    for key, value in row.items():
        if canvas_course_id in key:
            return 1 if parse_bool(value) else 0
    return None


def bronco_id_from_row(row):
    return row_value(row, "bid", "bronco_id", "broncoid", "bronco_number")


def person_row(conn, bronco_id):
    return conn.execute(
        """
        SELECT bronco_id, name, email, designation, active, notes, updated_at
        FROM people
        WHERE lower(bronco_id) = lower(?)
        """,
        (bronco_id,),
    ).fetchone()


def person_ref_for_bronco_id(bronco_id):
    normalized = str(bronco_id or "").strip().lower()
    if not normalized:
        return ""
    return hmac.new(
        PERSON_REF_SECRET.encode("utf-8"),
        normalized.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def person_row_from_ref(conn, person_ref):
    candidate = str(person_ref or "").strip()
    if not candidate:
        return None

    rows = conn.execute(
        """
        SELECT bronco_id, name, email, designation, active, notes, updated_at
        FROM people
        """
    ).fetchall()
    for row in rows:
        expected = person_ref_for_bronco_id(row["bronco_id"])
        if secrets.compare_digest(candidate, expected):
            return row
    return None


def people_rows(conn, include_bronco_id=False):
    rows = conn.execute(
        """
        SELECT bronco_id, name, email, designation, active, notes, updated_at
        FROM people
        WHERE NOT EXISTS (
            SELECT 1
            FROM cards
            WHERE cards.bronco_id != ''
              AND lower(cards.bronco_id) = lower(people.bronco_id)
        )
        ORDER BY
            CASE WHEN name = '' THEN 1 ELSE 0 END,
            name,
            bronco_id
        """
    ).fetchall()

    result = []
    for row in rows:
        person = {
            "person_ref": person_ref_for_bronco_id(row["bronco_id"]),
            "name": row["name"],
            "email": row["email"],
            "designation": row["designation"],
            "active": bool(row["active"]),
            "notes": row["notes"],
            "updated_at": row["updated_at"],
            "display_name": row["name"] or "Imported person",
        }
        if include_bronco_id:
            person["bronco_id"] = row["bronco_id"]
        result.append(person)

    return result

def sync_bronco_certifications_to_card(conn, bronco_id, card_id, actor=None):
    if not bronco_id or not card_id:
        return 0

    rows = conn.execute(
        """
        SELECT bronco_id, station_id, active, notes, updated_at
        FROM bronco_certifications
        WHERE bronco_id = ?
        """,
        (bronco_id,),
    ).fetchall()

    timestamp = now_iso()
    count = 0
    for row in rows:
        conn.execute(
            """
            INSERT INTO certifications (
                card_id,
                station_id,
                active,
                notes,
                updated_at,
                granted_via,
                granted_by,
                granted_at
            )
            VALUES (?, ?, ?, ?, ?, 'import', ?, ?)
            ON CONFLICT(card_id, station_id) DO UPDATE SET
                active = excluded.active,
                notes = CASE
                    WHEN certifications.notes = '' THEN excluded.notes
                    ELSE certifications.notes
                END,
                updated_at = excluded.updated_at,
                granted_via = CASE
                    WHEN excluded.active = 1 THEN excluded.granted_via
                    ELSE certifications.granted_via
                END,
                granted_by = CASE
                    WHEN excluded.active = 1 THEN excluded.granted_by
                    ELSE certifications.granted_by
                END,
                granted_at = CASE
                    WHEN excluded.active = 1 THEN excluded.granted_at
                    ELSE certifications.granted_at
                END
            """,
            (
                card_id,
                row["station_id"],
                row["active"],
                row["notes"],
                timestamp,
                account_field(actor, "card_id"),
                timestamp,
            ),
        )
        count += 1

    return count


def card_row(conn, card_id):
    return conn.execute(
        """
        SELECT card_id, bronco_id, name, email, designation, active, notes, updated_at
        FROM cards
        WHERE card_id = ?
        """,
        (card_id,),
    ).fetchone()


def card_display(row, fallback):
    if row and row["name"]:
        return row["name"]
    return fallback


def account_field(account, key, default=""):
    if not account:
        return default
    if isinstance(account, sqlite3.Row):
        return account[key] if key in account.keys() else default
    return account.get(key, default)


def add_audit_log(conn, account, action, target_type="", target_id="", details=None):
    conn.execute(
        """
        INSERT INTO audit_log (
            timestamp,
            actor_card_id,
            actor_name,
            actor_role,
            action,
            target_type,
            target_id,
            details
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            now_iso(),
            account_field(account, "card_id"),
            account_field(account, "name"),
            role_label(account_field(account, "role")),
            action,
            target_type,
            target_id,
            json.dumps(details or {}, sort_keys=True),
        ),
    )


def queue_canvas_sync_task(conn, card, station_id, active, actor=None):
    bronco_id = account_field(card, "bronco_id")
    if not bronco_id:
        return False

    conn.execute(
        """
        INSERT INTO canvas_sync_tasks (
            bronco_id,
            card_id,
            station_id,
            desired_active,
            status,
            created_at,
            created_by_card_id,
            created_by_name,
            completed_at,
            completed_by_card_id
        )
        VALUES (?, ?, ?, ?, 'pending', ?, ?, ?, '', '')
        ON CONFLICT(bronco_id, station_id) DO UPDATE SET
            card_id = excluded.card_id,
            desired_active = excluded.desired_active,
            status = 'pending',
            created_at = excluded.created_at,
            created_by_card_id = excluded.created_by_card_id,
            created_by_name = excluded.created_by_name,
            completed_at = '',
            completed_by_card_id = ''
        """,
        (
            bronco_id,
            account_field(card, "card_id"),
            station_id,
            1 if active else 0,
            now_iso(),
            account_field(actor, "card_id"),
            account_field(actor, "name"),
        ),
    )
    return True


def canvas_sync_task_rows(conn, include_bronco_id=False):
    rows = conn.execute(
        """
        SELECT
            canvas_sync_tasks.bronco_id,
            canvas_sync_tasks.card_id,
            canvas_sync_tasks.station_id,
            canvas_sync_tasks.desired_active,
            canvas_sync_tasks.created_at,
            canvas_sync_tasks.created_by_card_id,
            canvas_sync_tasks.created_by_name,
            people.name,
            people.email,
            stations.name AS station_name
        FROM canvas_sync_tasks
        JOIN people ON people.bronco_id = canvas_sync_tasks.bronco_id
        JOIN stations ON stations.id = canvas_sync_tasks.station_id
        WHERE canvas_sync_tasks.status = 'pending'
        ORDER BY canvas_sync_tasks.created_at, people.name, stations.name
        """
    ).fetchall()
    return [
        {
            "person_ref": person_ref_for_bronco_id(row["bronco_id"]),
            "card_id": row["card_id"],
            "station_id": row["station_id"],
            "station_name": row["station_name"],
            "desired_active": bool(row["desired_active"]),
            "created_at": row["created_at"],
            "created_by_card_id": row["created_by_card_id"],
            "created_by_name": row["created_by_name"],
            "name": row["name"],
            "email": row["email"],
            **({"bronco_id": row["bronco_id"]} if include_bronco_id else {}),
        }
        for row in rows
    ]


def card_can_enter(row):
    return bool(row and row["active"])


def certify_permission_rows(conn):
    rows = conn.execute(
        """
        SELECT card_id, station_id, updated_at
        FROM certify_permissions
        ORDER BY card_id, station_id
        """
    ).fetchall()
    return [dict(row) for row in rows]


def permitted_station_ids(conn, card_id):
    rows = conn.execute(
        """
        SELECT station_id
        FROM certify_permissions
        WHERE card_id = ?
        ORDER BY station_id
        """,
        (card_id,),
    ).fetchall()
    return [row["station_id"] for row in rows]


def account_can_certify_station(conn, account, station_id):
    if not account:
        return False

    role = normalize_access_role(account["role"])
    if role == "admin":
        return True

    if role not in ("staff", "volunteer"):
        return False

    row = conn.execute(
        """
        SELECT 1
        FROM certify_permissions
        WHERE card_id = ? AND station_id = ?
        """,
        (account["card_id"], station_id),
    ).fetchone()
    return bool(row)


def certifier_account_for_card(conn, card_id, station_id):
    row = conn.execute(
        """
        SELECT
            user_accounts.card_id,
            user_accounts.role,
            user_accounts.active,
            cards.name,
            cards.email,
            cards.active AS card_active
        FROM user_accounts
        JOIN cards ON cards.card_id = user_accounts.card_id
        WHERE user_accounts.card_id = ?
        """,
        (card_id,),
    ).fetchone()

    if not row or not row["active"] or not row["card_active"]:
        return None

    account = dict(row)
    if account_can_certify_station(conn, account, station_id):
        return account

    return None


def active_cert_mode(station_id):
    mode = cert_modes.get(station_id)
    if not mode:
        return None

    current_time = time.time()
    if (
        mode["arm_state"] == "pending_second_swipe"
        and mode["first_swipe_expires_at"] <= current_time
    ):
        cert_modes.pop(station_id, None)
        return None

    if mode["arm_state"] == "armed" and mode["expires_at"] <= current_time:
        cert_modes.pop(station_id, None)
        return None

    return mode


def start_pending_cert_mode(station_id, grantor, grantor_session_id=None):
    current_time = time.time()
    cert_modes[station_id] = {
        "grantor_card_id": grantor["card_id"],
        "grantor_name": grantor["name"] or grantor["card_id"],
        "grantor_role": grantor["role"],
        "expires_at": 0,
        "arm_state": "pending_second_swipe",
        "first_swipe_expires_at": current_time + CERT_CONFIRM_SECONDS,
        "grantor_session_id": grantor_session_id,
    }
    return cert_modes[station_id]


def arm_cert_mode(station_id, grantor, grantor_session_id=None):
    current_time = time.time()
    previous_mode = cert_modes.get(station_id)
    if (
        grantor_session_id is None
        and previous_mode
        and previous_mode.get("grantor_card_id") == grantor["card_id"]
    ):
        grantor_session_id = previous_mode.get("grantor_session_id")
    cert_modes[station_id] = {
        "grantor_card_id": grantor["card_id"],
        "grantor_name": grantor["name"] or grantor["card_id"],
        "grantor_role": grantor["role"],
        "expires_at": current_time + CERT_MODE_SECONDS,
        "arm_state": "armed",
        "first_swipe_expires_at": 0,
        "grantor_session_id": grantor_session_id,
    }
    return cert_modes[station_id]


def cert_mode_response(conn, card_id, station, action, led_signal, mode, event_id=""):
    result = log_event(
        conn,
        card_id,
        station["station_id"],
        station["station_name"],
        "station",
        action,
        details=f"Certification mode initiated by card {card_id}",
        event_id=event_id,
    )
    result["led_signal"] = led_signal
    result["cert_mode_expires_at"] = mode["expires_at"]
    return result


def is_certified(conn, card_id, station_id):
    row = conn.execute(
        """
        SELECT active
        FROM certifications
        WHERE card_id = ? AND station_id = ?
        """,
        (card_id, station_id),
    ).fetchone()
    return bool(row and row["active"])


def log_event(
    conn,
    card_id,
    station_id,
    station_name,
    station_kind,
    action,
    allowed=True,
    duration_seconds=None,
    warning="",
    details="",
    event_id="",
    event_time=None,
):
    event_time = event_time or now_iso()
    active_users = building_active_count(conn)

    conn.execute(
        """
        INSERT INTO swipe_events (
            card_id, station_id, station_name, station_kind, timestamp,
            action, allowed, duration_seconds, active_users, warning, details,
            event_id
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            card_id,
            station_id,
            station_name,
            station_kind,
            event_time,
            action,
            1 if allowed else 0,
            duration_seconds,
            active_users,
            warning,
            details,
            event_id,
        ),
    )

    led_signal = "access_granted" if allowed else "access_denied"

    return {
        "card_id": card_id,
        "station_id": station_id,
        "station_name": station_name,
        "station_kind": station_kind,
        "timestamp": event_time,
        "action": action,
        "allowed": allowed,
        "duration_seconds": duration_seconds,
        "active_users": active_users,
        "warning": warning,
        "details": details,
        "led_signal": led_signal,
    }


def logged_event_response(row):
    action = row["action"]
    allowed = bool(row["allowed"])
    if action == "cert_mode_pending":
        led_signal = "cert_mode_pending"
    elif action == "cert_mode_armed":
        led_signal = "cert_mode_armed"
    elif action == "certification_granted":
        led_signal = "cert_success"
    else:
        led_signal = "access_granted" if allowed else "access_denied"

    return {
        "card_id": row["card_id"],
        "station_id": row["station_id"],
        "station_name": row["station_name"],
        "station_kind": row["station_kind"],
        "timestamp": row["timestamp"],
        "action": action,
        "allowed": allowed,
        "duration_seconds": row["duration_seconds"],
        "active_users": row["active_users"],
        "warning": row["warning"],
        "details": row["details"],
        "led_signal": led_signal,
    }


def active_station_sessions_for_card(conn, card_id):
    return conn.execute(
        """
        SELECT id, card_id, station_id, station_name, started_at
        FROM active_sessions
        WHERE card_id = ?
        ORDER BY started_at
        """,
        (card_id,),
    ).fetchall()


def handle_door_swipe(conn, card_id, door_id, door_name, event_id=""):
    card = card_row(conn, card_id)
    active_person = conn.execute(
        """
        SELECT card_id, entered_at, entry_door_id, entry_door_name
        FROM active_people
        WHERE card_id = ?
        """,
        (card_id,),
    ).fetchone()

    if not active_person:
        if not card:
            return log_event(
                conn,
                card_id,
                door_id,
                door_name,
                "door",
                "denied",
                allowed=False,
                warning="unknown_card",
                details="Card is not in the card database",
                event_id=event_id,
            )

        if not card_can_enter(card):
            return log_event(
                conn,
                card_id,
                door_id,
                door_name,
                "door",
                "denied",
                allowed=False,
                warning="inactive_card",
                details="Card is disabled in the card database",
                event_id=event_id,
            )

        if not space_is_open() and not after_hours_access_role(conn, card_id):
            return log_event(
                conn,
                card_id,
                door_id,
                door_name,
                "door",
                "denied",
                allowed=False,
                warning="outside_open_hours",
                details=(
                    f"Regular user entry is limited to {opening_hours_label()} "
                    f"({SPACE_TIMEZONE_NAME})"
                ),
                event_id=event_id,
            )

        conn.execute(
            """
            INSERT INTO active_people (
                card_id, entered_at, entry_door_id, entry_door_name
            )
            VALUES (?, ?, ?, ?)
            """,
            (card_id, now_iso(), door_id, door_name),
        )
        return log_event(
            conn,
            card_id,
            door_id,
            door_name,
            "door",
            "enter",
            event_id=event_id,
        )

    exited_at = datetime.now(timezone.utc).replace(microsecond=0)
    entered_at = active_person["entered_at"]
    duration_seconds = elapsed_seconds(entered_at, exited_at)
    open_sessions = active_station_sessions_for_card(conn, card_id)

    conn.execute("DELETE FROM active_people WHERE card_id = ?", (card_id,))

    warning = ""
    details = ""

    if open_sessions:
        warning = "left_with_station_active"
        details = "; ".join(row["station_name"] for row in open_sessions)

        for session in open_sessions:
            started_at = session["started_at"]
            station_duration = elapsed_seconds(started_at, exited_at)
            conn.execute("DELETE FROM active_sessions WHERE id = ?", (session["id"],))
            log_event(
                conn,
                card_id,
                session["station_id"],
                session["station_name"],
                "station",
                "station_auto_out",
                duration_seconds=station_duration,
                warning=warning,
                details=f"Auto closed because card exited at {door_name}",
            )

    return log_event(
        conn,
        card_id,
        door_id,
        door_name,
        "door",
        "exit",
        duration_seconds=duration_seconds,
        warning=warning,
        details=details,
        event_id=event_id,
    )


def grant_certification_via_swipe(conn, card, station, mode, event_id=""):
    timestamp = now_iso()
    grantor_card_id = mode["grantor_card_id"]
    grantor_name = mode["grantor_name"]
    trainee_name = card_display(card, card["card_id"])

    conn.execute(
        """
        INSERT INTO certifications (
            card_id,
            station_id,
            active,
            notes,
            updated_at,
            granted_via,
            granted_by,
            granted_at
        )
        VALUES (?, ?, 1, ?, ?, 'swipe', ?, ?)
        ON CONFLICT(card_id, station_id) DO UPDATE SET
            active = 1,
            notes = CASE
                WHEN certifications.notes = '' THEN excluded.notes
                ELSE certifications.notes
            END,
            updated_at = excluded.updated_at,
            granted_via = excluded.granted_via,
            granted_by = excluded.granted_by,
            granted_at = excluded.granted_at
        """,
        (
            card["card_id"],
            station["station_id"],
            f"Swipe-certified by {grantor_name}",
            timestamp,
            grantor_card_id,
            timestamp,
        ),
    )

    if card["bronco_id"]:
        conn.execute(
            """
            INSERT INTO bronco_certifications (
                bronco_id, station_id, active, notes, updated_at
            )
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(bronco_id, station_id) DO UPDATE SET
                active = 1,
                notes = CASE
                    WHEN bronco_certifications.notes = '' THEN excluded.notes
                    ELSE bronco_certifications.notes
                END,
                updated_at = excluded.updated_at
            """,
            (
                card["bronco_id"],
                station["station_id"],
                f"Swipe-certified by {grantor_name}",
                timestamp,
            ),
        )
    grantor_account = {
        "card_id": grantor_card_id,
        "name": grantor_name,
        "role": mode.get("grantor_role", ""),
    }
    queue_canvas_sync_task(
        conn,
        card,
        station["station_id"],
        True,
        grantor_account,
    )
    add_audit_log(
        conn,
        grantor_account,
        "certification_granted_via_swipe",
        "certification",
        f"{card['card_id']}:{station['station_id']}",
        {
            "card_id": card["card_id"],
            "station_id": station["station_id"],
            "station_name": station["station_name"],
            "trainee_name": trainee_name,
        },
    )

    cert_modes.pop(station["station_id"], None)
    result = log_event(
        conn,
        card["card_id"],
        station["station_id"],
        station["station_name"],
        "station",
        "certification_granted",
        warning="cert_granted_via_swipe",
        details=(
            f"{trainee_name} certified for {station['station_name']} "
            f"via swipe by {grantor_name}"
        ),
        event_id=event_id,
    )
    result["led_signal"] = "cert_success"
    return result


def handle_station_swipe(conn, card_id, station_id, station, event_id=""):
    station_name = station["station_name"]
    requires_certification = station["requires_certification"]
    skip_cert_mode_for_this_swipe = False
    active_session = conn.execute(
        """
        SELECT id, started_at
        FROM active_sessions
        WHERE card_id = ? AND station_id = ?
        ORDER BY started_at DESC
        LIMIT 1
        """,
        (card_id, station_id),
    ).fetchone()

    mode = active_cert_mode(station_id) if requires_certification else None
    if mode and mode["arm_state"] == "pending_second_swipe":
        if card_id != mode["grantor_card_id"]:
            cert_modes.pop(station_id, None)
            mode = None
            skip_cert_mode_for_this_swipe = True
        elif (
            active_session
            and mode.get("grantor_session_id") == active_session["id"]
        ):
            grantor = certifier_account_for_card(conn, card_id, station_id)
            if grantor:
                mode = arm_cert_mode(
                    station_id,
                    grantor,
                    active_session["id"],
                )
                return cert_mode_response(
                    conn,
                    card_id,
                    station,
                    "cert_mode_armed",
                    "cert_mode_armed",
                    mode,
                    event_id,
                )
            cert_modes.pop(station_id, None)
            mode = None
    elif (
        mode
        and mode["arm_state"] == "armed"
        and card_id == mode["grantor_card_id"]
        and active_session
        and mode.get("grantor_session_id") == active_session["id"]
    ):
        grantor = certifier_account_for_card(conn, card_id, station_id)
        if grantor:
            mode = arm_cert_mode(
                station_id,
                grantor,
                active_session["id"],
            )
            return cert_mode_response(
                conn,
                card_id,
                station,
                "cert_mode_armed",
                "cert_mode_armed",
                mode,
                event_id,
            )
        cert_modes.pop(station_id, None)
        mode = None

    if active_session:
        ended_at = datetime.now(timezone.utc).replace(microsecond=0)
        started_at = active_session["started_at"]
        duration_seconds = elapsed_seconds(started_at, ended_at)

        conn.execute("DELETE FROM active_sessions WHERE id = ?", (active_session["id"],))
        warning, details = station_override_warning(station)
        return log_event(
            conn,
            card_id,
            station_id,
            station_name,
            "station",
            "station_out",
            duration_seconds=duration_seconds,
            warning=warning,
            details=details,
            event_id=event_id,
        )

    if requires_certification:
        mode = active_cert_mode(station_id)
        if mode and mode["arm_state"] == "pending_second_swipe":
            if card_id == mode["grantor_card_id"]:
                grantor = certifier_account_for_card(conn, card_id, station_id)
                if grantor:
                    mode = arm_cert_mode(station_id, grantor)
                    return cert_mode_response(
                        conn,
                        card_id,
                        station,
                        "cert_mode_armed",
                        "cert_mode_armed",
                        mode,
                        event_id,
                    )
                cert_modes.pop(station_id, None)
            else:
                cert_modes.pop(station_id, None)
                skip_cert_mode_for_this_swipe = True
    else:
        cert_modes.pop(station_id, None)

    card = card_row(conn, card_id)
    if not card:
        return log_event(
            conn,
            card_id,
            station_id,
            station_name,
            "station",
            "denied",
            allowed=False,
            warning="unknown_card",
            details="Card is not in the card database",
            event_id=event_id,
        )

    if not card_can_enter(card):
        return log_event(
            conn,
            card_id,
            station_id,
            station_name,
            "station",
            "denied",
            allowed=False,
            warning="inactive_card",
            details="Card is disabled in the card database",
            event_id=event_id,
        )

    grantor = (
        certifier_account_for_card(conn, card_id, station_id)
        if requires_certification
        else None
    )
    mode = active_cert_mode(station_id) if requires_certification else None

    if requires_certification and not skip_cert_mode_for_this_swipe:
        if mode and mode["arm_state"] == "armed":
            if grantor:
                mode = arm_cert_mode(station_id, grantor)
                return cert_mode_response(
                    conn,
                    card_id,
                    station,
                    "cert_mode_armed",
                    "cert_mode_armed",
                    mode,
                    event_id,
                )
            return grant_certification_via_swipe(
                conn,
                card,
                station,
                mode,
                event_id,
            )

    override_active = override_is_effective(station)

    if (
        requires_certification
        and not override_active
        and not grantor
        and not is_certified(conn, card_id, station_id)
    ):
        return log_event(
            conn,
            card_id,
            station_id,
            station_name,
            "station",
            "denied",
            allowed=False,
            warning="not_certified",
            details="Card is not certified for this station",
            event_id=event_id,
        )

    moved_from = active_station_sessions_for_card(conn, card_id)
    if moved_from:
        ended_at = datetime.now(timezone.utc).replace(microsecond=0)
        for session in moved_from:
            started_at = session["started_at"]
            station_duration = elapsed_seconds(started_at, ended_at)
            conn.execute("DELETE FROM active_sessions WHERE id = ?", (session["id"],))
            log_event(
                conn,
                card_id,
                session["station_id"],
                session["station_name"],
                "station",
                "station_auto_out",
                duration_seconds=station_duration,
                warning="moved_station_without_swipe_out",
                details=f"Auto closed before starting {station_name}",
            )

    active_person = conn.execute(
        "SELECT card_id FROM active_people WHERE card_id = ?",
        (card_id,),
    ).fetchone()

    warning = ""
    details = ""
    if moved_from:
        warning = "moved_station_without_swipe_out"
        details = "; ".join(row["station_name"] for row in moved_from)
    elif not active_person:
        warning = "station_swipe_without_door_entry"
        details = "Card is not currently counted inside"

    warning, details = station_override_warning(station, warning, details)

    session_cursor = conn.execute(
        """
        INSERT INTO active_sessions (
            card_id, station_id, station_name, started_at
        )
        VALUES (?, ?, ?, ?)
        """,
        (card_id, station_id, station_name, now_iso()),
    )

    result = log_event(
        conn,
        card_id,
        station_id,
        station_name,
        "station",
        "station_in",
        warning=warning,
        details=details,
        event_id=event_id,
    )
    if (
        requires_certification
        and grantor
        and not skip_cert_mode_for_this_swipe
    ):
        mode = start_pending_cert_mode(
            station_id,
            grantor,
            session_cursor.lastrowid,
        )
        result["led_signal"] = "cert_mode_pending"
        result["cert_mode_expires_at"] = mode["first_swipe_expires_at"]
    return result


@app.post("/swipe")
def swipe():
    key_error = require_station_api_key()
    if key_error:
        return key_error

    data = request_json()
    try:
        card_id = validate_text(
            data.get("card_id"),
            "card_id",
            MAX_ID_LENGTH,
            required=True,
        )
        station_id = validate_text(
            data.get("station_id"),
            "station_id",
            MAX_ID_LENGTH,
            required=True,
        )
        station_name = validate_text(
            data.get("station_name"),
            "station_name",
            MAX_NAME_LENGTH,
        )
        event_id = validate_text(
            data.get("event_id"),
            "event_id",
            MAX_ID_LENGTH,
        )
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400

    station_kind = str(data.get("station_kind", data.get("kind", ""))).strip()

    with db_lock:
        conn = db_connect()
        if event_id:
            existing = conn.execute(
                """
                SELECT
                    card_id,
                    station_id,
                    station_name,
                    station_kind,
                    timestamp,
                    action,
                    allowed,
                    duration_seconds,
                    active_users,
                    warning,
                    details
                FROM swipe_events
                WHERE event_id = ?
                """,
                (event_id,),
            ).fetchone()
            if existing:
                conn.close()
                return jsonify(
                    {
                        "ok": True,
                        "duplicate": True,
                        **logged_event_response(existing),
                    }
                )
        station = get_station_info(
            conn,
            station_id,
            station_name,
            station_kind,
        )

        if station["station_kind"] == "door":
            result = handle_door_swipe(
                conn,
                card_id,
                station_id,
                station["station_name"],
                event_id,
            )
        else:
            result = handle_station_swipe(
                conn,
                card_id,
                station_id,
                station,
                event_id,
            )

        conn.commit()
        conn.close()

    return jsonify({"ok": True, **result})


@app.post("/api/test-swipe")
def test_swipe():
    """Admin-only hardware-free swipe simulator for dashboard testing."""
    role, error = require_access("admin")
    if error:
        return error

    data = request_json()
    try:
        card_id = validate_text(
            data.get("card_id"),
            "card_id",
            MAX_ID_LENGTH,
            required=True,
        )
        station_id = validate_text(
            data.get("station_id"),
            "station_id",
            MAX_ID_LENGTH,
            required=True,
        )
    except ValueError:
        return jsonify({"ok": False, "error": "Choose both a card and a reader"}), 400

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        configured_station = conn.execute(
            "SELECT id FROM stations WHERE id = ?",
            (station_id,),
        ).fetchone()
        if not configured_station:
            conn.close()
            return jsonify({"ok": False, "error": "Reader location not found"}), 404
        station = get_station_info(conn, station_id)
        if station["station_kind"] == "door":
            result = handle_door_swipe(conn, card_id, station_id, station["station_name"])
        else:
            result = handle_station_swipe(conn, card_id, station_id, station)
        add_audit_log(conn, account, "test_swipe", "station", station_id, {"card_id": card_id})
        conn.commit()
        conn.close()

    return jsonify({"ok": True, "simulated": True, **result})


@app.get("/health")
def health():
    try:
        with db_lock:
            conn = db_connect()
            conn.execute("SELECT 1").fetchone()
            conn.close()
    except sqlite3.Error:
        app.logger.exception("Database health check failed")
        return jsonify({"ok": False, "time": now_iso(), "database": "error"}), 503

    return jsonify(
        {
            "ok": True,
            "time": now_iso(),
            "database": "ok",
            "station_api_key_configured": bool(STATION_API_KEY),
        }
    )


@app.get("/")
@app.get("/dashboard")
def dashboard():
    with open(DASHBOARD_PATH, "r", encoding="utf-8") as file:
        return Response(file.read(), mimetype="text/html")


@app.get("/api/dashboard")
def dashboard_data():
    limit = request.args.get("limit", "20")
    swipe_query = request.args.get("swipe_query", "").strip()[:100]
    swipe_station = request.args.get("swipe_station", "").strip()[:MAX_ID_LENGTH]
    swipe_warning_only = request.args.get("swipe_warning_only", "") == "1"
    account = account_for_token(token_from_request())
    role = normalize_access_role(account["role"]) if account else ""
    can_view_people = role in ("admin", "staff", "volunteer")
    can_view_full_dashboard = role in ("admin", "staff")

    try:
        limit = int(limit)
    except ValueError:
        limit = 20

    limit = max(1, min(limit, 100))

    with db_lock:
        conn = db_connect()
        total_active = building_active_count(conn)
        station_status = station_status_rows(conn)
        active_people = active_people_rows(conn) if can_view_people else []
        stations = [
            {
                **station,
                "active_cards": (
                    station["active_cards"]
                    if can_view_full_dashboard
                    else ""
                ),
            }
            for station in station_status
        ]
        active_sessions = active_session_rows(conn) if can_view_full_dashboard else []
        recent_swipes = (
            recent_swipe_rows(conn, limit, swipe_query, swipe_station, swipe_warning_only)
            if can_view_full_dashboard
            else []
        )
        latest_swipes = recent_swipe_rows(conn, 1)
        warnings = warning_rows(conn, 10) if can_view_full_dashboard else []
        conn.close()

    return jsonify(
        {
            "server_time": now_iso(),
            "dashboard_access": role or "public",
            "total_active": total_active,
            "active_station_count": len(
                [
                    station
                    for station in station_status
                    if station["active_sessions"] > 0
                ]
            ),
            "last_swipe_at": (
                latest_swipes[0]["timestamp"]
                if latest_swipes
                else ""
            ),
            "stations": stations,
            "active_people": active_people,
            "active_sessions": active_sessions,
            "recent_swipes": recent_swipes,
            "warnings": warnings,
        }
    )


@app.post("/api/access")
def access_check():
    if request.content_length and request.content_length > MAX_LOGIN_BODY_BYTES:
        return jsonify({"ok": False, "error": "Login request is too large"}), 413

    data = request_json()
    login = str(data.get("login", "")).strip().lower()
    password = str(data.get("password", ""))

    prune_in_memory_auth_state()

    if not login or not password:
        return jsonify({"ok": False, "error": "Login and password are required"}), 400
    if len(login) > MAX_EMAIL_LENGTH or len(password) > 1024:
        return jsonify({"ok": False, "error": "Login not recognized"}), 403
    retry_after = login_retry_after(login)
    if retry_after:
        if should_log_security_event("login_rate_limited", login):
            with db_lock:
                conn = db_connect()
                add_audit_log(
                    conn,
                    None,
                    "login_rate_limited",
                    "account",
                    login,
                    {
                        "login": login,
                        "source_ip": client_ip_address(),
                        "retry_after_seconds": retry_after,
                    },
                )
                conn.commit()
                conn.close()
        response = jsonify(
            {
                "ok": False,
                "error": "Too many failed attempts; try again later",
                "retry_after_seconds": retry_after,
            }
        )
        response.headers["Retry-After"] = str(retry_after)
        return response, 429

    with db_lock:
        conn = db_connect()
        account = conn.execute(
            """
            SELECT
                user_accounts.card_id,
                user_accounts.username,
                user_accounts.role,
                user_accounts.password_hash,
                user_accounts.active,
                cards.name,
                cards.email
            FROM user_accounts
            JOIN cards ON cards.card_id = user_accounts.card_id
            WHERE lower(user_accounts.username) = lower(?)
            LIMIT 1
            """,
            (login,),
        ).fetchone()

        password_hash = account["password_hash"] if account else DUMMY_PASSWORD_HASH
        password_valid = check_password_hash(password_hash, password)
        account_valid = bool(
            account
            and account["active"]
            and normalize_access_role(account["role"])
            and password_valid
        )
        if not account_valid:
            record_login_failure(login)
            retry_after = login_retry_after(login)
            if should_log_security_event("login_failed", login):
                add_audit_log(
                    conn,
                    None,
                    "login_failed",
                    "account",
                    login,
                    {
                        "login": login,
                        "source_ip": client_ip_address(),
                    },
                )
            conn.commit()
            conn.close()
            if retry_after:
                response = jsonify(
                    {
                        "ok": False,
                        "error": "Too many failed attempts; try again later",
                        "retry_after_seconds": retry_after,
                    }
                )
                response.headers["Retry-After"] = str(retry_after)
                return response, 429
            return jsonify({"ok": False, "error": "Login not recognized"}), 403

        account = dict(account)
        token = secrets.token_urlsafe(32)
        role = normalize_access_role(account["role"])
        clear_login_failures(login)
        with session_lock:
            account_sessions = sorted(
                (
                    (existing_token, session)
                    for existing_token, session in session_tokens.items()
                    if session.get("card_id") == account["card_id"]
                ),
                key=lambda item: item[1].get("created_at", 0),
            )
            while len(account_sessions) >= SESSION_LIMIT_PER_ACCOUNT:
                oldest_token, _ = account_sessions.pop(0)
                session_tokens.pop(oldest_token, None)
            created_at = time.time()
            session_tokens[token] = {
                "card_id": account["card_id"],
                "created_at": created_at,
                "expires_at": created_at + SESSION_MAX_SECONDS,
            }
        add_audit_log(
            conn,
            account,
            "login",
            "account",
            account["card_id"],
            {"login": login, "source_ip": client_ip_address()},
        )
        conn.commit()
        conn.close()

    return jsonify(
        {
            "ok": True,
            "access_token": token,
            "role": role,
            "card_id": account["card_id"],
            "username": account["username"],
            "name": account["name"],
            "email": account["email"],
        }
    )


@app.post("/api/logout")
def logout():
    token = token_from_request()
    account = account_for_token(token)
    with db_lock:
        conn = db_connect()
        if account:
            add_audit_log(conn, account, "logout", "account", account["card_id"])
        conn.commit()
        conn.close()
    with session_lock:
        session_tokens.pop(token, None)
    return jsonify({"ok": True})


@app.get("/api/admin")
def admin_data():
    role, error = require_access("staff", "volunteer")
    if error:
        return error

    account = account_for_token(token_from_request())

    with db_lock:
        conn = db_connect()
        include_sensitive = role == "admin"
        cards = card_rows(
            conn,
            include_accounts=include_sensitive,
            include_bronco_id=include_sensitive,
        )
        people = people_rows(conn, include_bronco_id=include_sensitive)
        stations = station_option_rows(conn)
        readers = reader_option_rows(conn)
        certifications = certification_rows(
            conn,
            include_bronco_id=include_sensitive,
        )
        certify_permissions = certify_permission_rows(conn) if role == "admin" else []
        canvas_sync_tasks = (
            canvas_sync_task_rows(conn, include_bronco_id=(role == "admin"))
            if role in ("admin", "staff")
            else []
        )
        audit_log = audit_log_rows(conn) if role == "admin" else []
        can_certify_station_ids = (
            [station["station_id"] for station in stations]
            if role == "admin"
            else permitted_station_ids(conn, account["card_id"])
        )
        conn.close()

    return jsonify(
        {
            "role": role,
            "cards": cards,
            "people": people,
            "stations": stations,
            "readers": readers,
            "certifications": certifications,
            "certify_permissions": certify_permissions,
            "canvas_sync_tasks": canvas_sync_tasks,
            "audit_log": audit_log,
            "can_certify_station_ids": can_certify_station_ids,
        }
    )


@app.get("/api/pending-cards")
def pending_cards():
    role, error = require_access("staff")
    if error:
        return error

    with db_lock:
        conn = db_connect()
        cutoff = (
            datetime.now(timezone.utc).replace(microsecond=0)
            - timedelta(days=PENDING_CARD_DAYS)
        ).isoformat()
        rows = conn.execute(
            """
            SELECT
                latest.card_id,
                latest.timestamp AS last_seen_at,
                COALESCE(
                    NULLIF(latest.station_name, ''),
                    latest.station_id
                ) AS last_seen_station,
                COUNT(*) OVER () AS pending_count
            FROM swipe_events AS latest
            JOIN (
                SELECT card_id, MAX(id) AS latest_id
                FROM swipe_events
                WHERE warning = 'unknown_card' AND timestamp >= ?
                GROUP BY card_id
            ) AS newest
                ON newest.latest_id = latest.id
            LEFT JOIN cards
                ON cards.card_id = latest.card_id
            LEFT JOIN pending_card_dismissals AS dismissals
                ON dismissals.card_id = latest.card_id
            WHERE cards.card_id IS NULL
              AND (
                  dismissals.card_id IS NULL
                  OR latest.id > dismissals.ignored_through_event_id
              )
            ORDER BY latest.id DESC
            LIMIT 100
            """,
            (cutoff,),
        ).fetchall()
        conn.close()

    return jsonify(
        {
            "pending_count": int(rows[0]["pending_count"]) if rows else 0,
            "pending_cards": [
                {
                    "card_id": row["card_id"],
                    "last_seen_at": row["last_seen_at"],
                    "last_seen_station": row["last_seen_station"],
                }
                for row in rows
            ]
        }
    )


@app.post("/api/import/canvas")
def import_canvas_csv():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    try:
        rows = normalized_csv_rows(csv_text_from_request())
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400
    if not rows:
        return jsonify({"ok": False, "error": "CSV has no rows"}), 400

    base_columns = (
        ("Email", ("email", "email_address")),
        ("Student", ("student", "name", "person", "full_name")),
        ("bid", ("bid", "bronco_id", "broncoid", "bronco_number")),
    )
    missing_base_columns = [
        label
        for label, alternatives in base_columns
        if not any(alternative in rows[0] for alternative in alternatives)
    ]
    if missing_base_columns:
        return jsonify(
            {
                "ok": False,
                "error": "Canvas CSV is missing required columns: "
                + ", ".join(missing_base_columns),
            }
        ), 400

    missing_columns = [
        heading
        for heading, course_id, _ in CERTIFICATION_IMPORT_COLUMNS
        if not any(course_id in key for key in rows[0])
    ]
    if missing_columns:
        return jsonify(
            {
                "ok": False,
                "error": "Canvas CSV is missing certification columns",
                "missing_columns": missing_columns,
            }
        ), 400

    imported_rows = 0
    people_created = 0
    people_updated = 0
    certification_statuses = 0
    synced_certifications = 0
    pending_preserved = 0
    pending_resolved = 0
    errors = []
    seen_bronco_ids = {}

    with db_lock:
        conn = db_connect()
        station_ids = {
            row["id"]
            for row in conn.execute(
                "SELECT id FROM stations WHERE kind = 'station'"
            ).fetchall()
        }
        missing_stations = [
            station_id
            for _, _, station_id in CERTIFICATION_IMPORT_COLUMNS
            if station_id not in station_ids
        ]
        if missing_stations:
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "Server station configuration is missing: "
                    + ", ".join(missing_stations),
                }
            ), 500

        timestamp = now_iso()
        for line_number, row in enumerate(rows, start=2):
            try:
                bronco_id = validate_text(
                    bronco_id_from_row(row),
                    "bid",
                    MAX_ID_LENGTH,
                    required=True,
                )
                name = validate_text(
                    student_name_from_row(row),
                    "Student",
                    MAX_NAME_LENGTH,
                    required=True,
                )
                email = validate_email(
                    row_value(row, "email", "email_address"),
                    required=True,
                )
            except ValueError as error:
                errors.append({"line": line_number, "error": str(error)})
                continue

            bronco_key = bronco_id.lower()
            if bronco_key in seen_bronco_ids:
                errors.append(
                    {
                        "line": line_number,
                        "bronco_id": bronco_id,
                        "error": f"duplicate bid also appears on line {seen_bronco_ids[bronco_key]}",
                    }
                )
                continue
            seen_bronco_ids[bronco_key] = line_number

            existing = person_row(conn, bronco_id)
            stored_bronco_id = existing["bronco_id"] if existing else bronco_id
            if existing:
                conn.execute(
                    """
                    UPDATE people
                    SET name = ?, email = ?, active = 1, updated_at = ?
                    WHERE bronco_id = ?
                    """,
                    (name, email, timestamp, stored_bronco_id),
                )
                people_updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO people (
                        bronco_id, name, email, designation, active, notes, updated_at
                    )
                    VALUES (?, ?, ?, 'User', 1, 'Imported from Canvas', ?)
                    """,
                    (stored_bronco_id, name, email, timestamp),
                )
                people_created += 1

            conn.execute(
                """
                UPDATE cards
                SET name = ?, email = ?, updated_at = ?
                WHERE bronco_id = ?
                """,
                (name, email, timestamp, stored_bronco_id),
            )
            assigned_cards = conn.execute(
                "SELECT card_id FROM cards WHERE bronco_id = ?",
                (stored_bronco_id,),
            ).fetchall()

            for _, course_id, station_id in CERTIFICATION_IMPORT_COLUMNS:
                canvas_active = certification_import_value(row, course_id)
                pending = conn.execute(
                    """
                    SELECT desired_active
                    FROM canvas_sync_tasks
                    WHERE bronco_id = ? AND station_id = ? AND status = 'pending'
                    """,
                    (stored_bronco_id, station_id),
                ).fetchone()

                if pending and int(pending["desired_active"]) != canvas_active:
                    pending_preserved += 1
                    continue
                if pending:
                    conn.execute(
                        """
                        UPDATE canvas_sync_tasks
                        SET status = 'completed',
                            completed_at = ?,
                            completed_by_card_id = 'canvas-import'
                        WHERE bronco_id = ? AND station_id = ?
                        """,
                        (timestamp, stored_bronco_id, station_id),
                    )
                    pending_resolved += 1

                if canvas_active:
                    conn.execute(
                        """
                        INSERT INTO bronco_certifications (
                            bronco_id, station_id, active, notes, updated_at
                        )
                        VALUES (?, ?, 1, 'Imported from Canvas', ?)
                        ON CONFLICT(bronco_id, station_id) DO UPDATE SET
                            active = 1,
                            notes = excluded.notes,
                            updated_at = excluded.updated_at
                        """,
                        (stored_bronco_id, station_id, timestamp),
                    )
                else:
                    conn.execute(
                        """
                        DELETE FROM bronco_certifications
                        WHERE bronco_id = ? AND station_id = ?
                        """,
                        (stored_bronco_id, station_id),
                    )
                    conn.execute(
                        """
                        DELETE FROM certifications
                        WHERE station_id = ?
                          AND card_id IN (
                              SELECT card_id FROM cards WHERE bronco_id = ?
                          )
                        """,
                        (station_id, stored_bronco_id),
                    )
                certification_statuses += 1

            for card in assigned_cards:
                synced_certifications += sync_bronco_certifications_to_card(
                    conn, stored_bronco_id, card["card_id"], account
                )
            imported_rows += 1

        add_audit_log(
            conn,
            account,
            "canvas_database_imported",
            "canvas",
            "csv",
            {
                "rows": imported_rows,
                "people_created": people_created,
                "people_updated": people_updated,
                "certification_statuses": certification_statuses,
                "pending_preserved": pending_preserved,
                "pending_resolved": pending_resolved,
                "errors": len(errors),
            },
        )
        conn.commit()
        people = people_rows(conn, include_bronco_id=True)
        certifications = certification_rows(conn, include_bronco_id=True)
        canvas_sync_tasks = canvas_sync_task_rows(conn, include_bronco_id=True)
        conn.close()

    status = 200 if imported_rows else 400
    return jsonify(
        {
            "ok": imported_rows > 0,
            "imported_rows": imported_rows,
            "people_created": people_created,
            "people_updated": people_updated,
            "certification_statuses": certification_statuses,
            "synced_certifications": synced_certifications,
            "pending_preserved": pending_preserved,
            "pending_resolved": pending_resolved,
            "errors": errors[:25],
            "error_count": len(errors),
            "people": people,
            "certifications": certifications,
            "canvas_sync_tasks": canvas_sync_tasks,
        }
    ), status


@app.post("/api/canvas-sync-tasks/complete")
def complete_canvas_sync_task():
    role, error = require_access("staff")
    if error:
        return error

    account = account_for_token(token_from_request())
    data = request_json()
    try:
        person_ref = validate_text(
            data.get("person_ref"),
            "person_ref",
            128,
            required=True,
        )
        station_id = validate_text(
            data.get("station_id"),
            "station_id",
            MAX_ID_LENGTH,
            required=True,
        )
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400

    with db_lock:
        conn = db_connect()
        person = person_row_from_ref(conn, person_ref)
        if not person:
            conn.close()
            return jsonify({"ok": False, "error": "Pending Canvas task person not found"}), 404
        bronco_id = person["bronco_id"]
        task = conn.execute(
            """
            SELECT desired_active
            FROM canvas_sync_tasks
            WHERE bronco_id = ? AND station_id = ? AND status = 'pending'
            """,
            (bronco_id, station_id),
        ).fetchone()
        if not task:
            conn.close()
            return jsonify({"ok": False, "error": "Pending Canvas task not found"}), 404

        conn.execute(
            """
            UPDATE canvas_sync_tasks
            SET status = 'completed', completed_at = ?, completed_by_card_id = ?
            WHERE bronco_id = ? AND station_id = ?
            """,
            (now_iso(), account["card_id"], bronco_id, station_id),
        )
        add_audit_log(
            conn,
            account,
            "canvas_sync_task_completed",
            "canvas_certification",
            f"{bronco_id}:{station_id}",
            {"desired_active": bool(task["desired_active"])},
        )
        conn.commit()
        tasks = canvas_sync_task_rows(conn, include_bronco_id=(role == "admin"))
        conn.close()

    return jsonify({"ok": True, "canvas_sync_tasks": tasks})


@app.post("/api/cards")
def save_card():
    role, error = require_access("staff")
    if error:
        return error

    account = account_for_token(token_from_request())
    data = request_json()
    try:
        card_id = validate_text(
            data.get("card_id"),
            "card_id",
            MAX_ID_LENGTH,
            required=True,
        )
        provided_bronco_id = validate_text(
            data.get("bronco_id"),
            "BroncoID",
            MAX_ID_LENGTH,
        )
        person_ref = validate_text(
            data.get("person_ref"),
            "person_ref",
            128,
        )
        name = validate_text(data.get("name"), "name", MAX_NAME_LENGTH)
        email = validate_email(data.get("email"))
        notes = validate_text(data.get("notes"), "notes", MAX_NOTES_LENGTH)
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400

    designation_value = str(data.get("designation", "User") or "User").strip()
    if designation_value.lower() not in ("staff", "volunteer", "user"):
        return jsonify({"ok": False, "error": "designation is invalid"}), 400
    designation = normalize_designation(designation_value)
    active = 1 if parse_bool(data.get("active", True)) else 0
    account_fields_present = any(
        key in data
        for key in ("login_role", "login_password", "login_active")
    )

    if account_fields_present and role != "admin":
        return jsonify(
            {
                "ok": False,
                "error": "Admin login required to edit dashboard credentials",
            }
        ), 403

    with db_lock:
        conn = db_connect()
        existing_card = card_row(conn, card_id)
        target_account = conn.execute(
            """
            SELECT card_id, username, role, password_hash, active
            FROM user_accounts
            WHERE card_id = ?
            """,
            (card_id,),
        ).fetchone()

        if target_account and role != "admin":
            actor_level = ROLE_LEVELS.get(role, 0)
            target_level = ROLE_LEVELS.get(
                normalize_access_role(target_account["role"]),
                0,
            )
            if target_level > actor_level:
                conn.close()
                return jsonify(
                    {
                        "ok": False,
                        "error": "You cannot edit a higher-access account's card",
                    }
                ), 403

        linked_person = person_row_from_ref(conn, person_ref) if person_ref else None

        if person_ref and not linked_person:
            conn.close()
            return jsonify(
                {"ok": False, "error": "Imported person selection expired; select the person again"}
            ), 400

        if existing_card:
            bronco_id = existing_card["bronco_id"]
            if bronco_id:
                if (
                    linked_person
                    and linked_person["bronco_id"].lower() != bronco_id.lower()
                ):
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "A saved card cannot be reassigned to another person",
                        }
                    ), 409
                if (
                    provided_bronco_id
                    and provided_bronco_id.lower() != bronco_id.lower()
                ):
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "A saved card's BroncoID cannot be changed",
                        }
                    ), 409
                linked_person = person_row(conn, bronco_id)
            else:
                if role != "admin":
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Admin login required to repair this legacy card",
                        }
                    ), 403
                if linked_person:
                    bronco_id = linked_person["bronco_id"]
                elif provided_bronco_id:
                    linked_person = person_row(conn, provided_bronco_id)
                    bronco_id = (
                        linked_person["bronco_id"]
                        if linked_person
                        else provided_bronco_id
                    )
                else:
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Assign an imported person or enter a BroncoID to repair this legacy card",
                        }
                    ), 400
        elif linked_person:
            bronco_id = linked_person["bronco_id"]
            if provided_bronco_id and provided_bronco_id.lower() != bronco_id.lower():
                conn.close()
                return jsonify(
                    {"ok": False, "error": "BroncoID does not match the imported person"}
                ), 400
        elif provided_bronco_id:
            if person_row(conn, provided_bronco_id):
                conn.close()
                return jsonify(
                    {
                        "ok": False,
                        "error": "This person is already imported; select them using Assign Person",
                    }
                ), 409
            bronco_id = provided_bronco_id
        else:
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "BroncoID is required unless an imported person is selected",
                }
            ), 400

        duplicate_card = conn.execute(
            """
            SELECT card_id
            FROM cards
            WHERE lower(bronco_id) = lower(?) AND card_id != ?
            LIMIT 1
            """,
            (bronco_id, card_id),
        ).fetchone()
        if duplicate_card:
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": f"This person is already assigned to card {duplicate_card['card_id']}",
                }
            ), 409

        if linked_person:
            if not existing_card:
                name = linked_person["name"]
                email = linked_person["email"]
                notes = notes or linked_person["notes"]
                active = 1 if linked_person["active"] else 0

        if not name:
            conn.close()
            return jsonify({"ok": False, "error": "name is required"}), 400

        if role == "volunteer" and designation == "Staff":
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "Volunteers cannot designate cards as Staff",
                }
            ), 403

        if (
            role == "volunteer"
            and existing_card
            and normalize_designation(existing_card["designation"]) == "Staff"
        ):
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "Volunteers cannot edit Staff cards",
                }
            ), 403

        created_person = False
        if not linked_person:
            timestamp = now_iso()
            conn.execute(
                """
                INSERT INTO people (
                    bronco_id, name, email, designation, active, notes, updated_at
                )
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                (bronco_id, name, email, designation, notes, timestamp),
            )
            linked_person = person_row(conn, bronco_id)
            created_person = True
        else:
            conn.execute(
                """
                UPDATE people
                SET
                    name = ?,
                    email = ?,
                    designation = ?,
                    notes = ?,
                    updated_at = ?
                WHERE bronco_id = ?
                """,
                (
                    name,
                    email,
                    designation,
                    notes,
                    now_iso(),
                    bronco_id,
                ),
            )

        conn.execute(
            """
            INSERT INTO cards (
                card_id, bronco_id, name, email, designation, active, notes, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(card_id) DO UPDATE SET
                bronco_id = excluded.bronco_id,
                name = excluded.name,
                email = excluded.email,
                designation = excluded.designation,
                active = excluded.active,
                notes = excluded.notes,
                updated_at = excluded.updated_at
            """,
            (card_id, bronco_id, name, email, designation, active, notes, now_iso()),
        )

        synced_certifications = sync_bronco_certifications_to_card(
            conn, bronco_id, card_id, account
        )
        conn.execute(
            "DELETE FROM pending_card_dismissals WHERE card_id = ?",
            (card_id,),
        )

        login_role = normalize_access_role(target_account["role"]) if target_account else ""
        login_active = bool(target_account["active"]) if target_account else False
        password_changed = False
        if account_fields_present:
            raw_login_role = str(data.get("login_role", "") or "").strip()
            login_role = normalize_access_role(raw_login_role)
            login_password = str(data.get("login_password", ""))
            login_active = 1 if parse_bool(data.get("login_active", True)) else 0

            if raw_login_role and not login_role:
                conn.close()
                return jsonify({"ok": False, "error": "login role is invalid"}), 400
            if len(login_password) > 1024:
                conn.close()
                return jsonify(
                    {"ok": False, "error": "password is too long"}
                ), 400
            if account and account["card_id"] == card_id and (
                login_role != "admin" or not login_active
            ):
                conn.close()
                return jsonify(
                    {
                        "ok": False,
                        "error": "You cannot remove, demote, or disable the account you are signed in with",
                    }
                ), 409

            if not login_role:
                conn.execute(
                    "DELETE FROM certify_permissions WHERE card_id = ?",
                    (card_id,),
                )
                conn.execute("DELETE FROM user_accounts WHERE card_id = ?", (card_id,))
            else:
                login_username = login_username_from_email(email)
                if not login_username and target_account:
                    login_username = target_account["username"]
                if not login_username:
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "An email address is required to create a BroncoName login",
                        }
                    ), 400
                duplicate = conn.execute(
                    """
                    SELECT card_id
                    FROM user_accounts
                    WHERE lower(username) = lower(?) AND card_id != ?
                    """,
                    (login_username, card_id),
                ).fetchone()
                if duplicate:
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": f"Login username '{login_username}' is already used",
                        }
                    ), 400

                if not login_password and not target_account:
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Password is required when creating a dashboard login",
                        }
                    ), 400
                if login_password and (
                    len(login_password) < 8 or not login_password.strip()
                ):
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Password must be at least 8 characters",
                        }
                    ), 400

                password_hash = (
                    generate_password_hash(login_password)
                    if login_password
                    else target_account["password_hash"]
                )
                password_changed = bool(login_password)
                conn.execute(
                    """
                    INSERT INTO user_accounts (
                        card_id, username, role, password_hash, active, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(card_id) DO UPDATE SET
                        username = excluded.username,
                        role = excluded.role,
                        password_hash = excluded.password_hash,
                        active = excluded.active,
                        updated_at = excluded.updated_at
                    """,
                    (
                        card_id,
                        login_username,
                        role_label(login_role),
                        password_hash,
                        login_active,
                        now_iso(),
                    ),
                )
                if login_role == "admin":
                    conn.execute(
                        "DELETE FROM certify_permissions WHERE card_id = ?",
                        (card_id,),
                    )

        add_audit_log(
            conn,
            account,
            "card_updated" if existing_card else "card_created",
            "card",
            card_id,
            {
                "bronco_id": bronco_id,
                "name": name,
                "email": email,
                "designation": designation,
                "active": bool(active),
                "person_created": created_person,
                "selected_from_import": bool(person_ref),
                "login_fields_changed": account_fields_present,
                "login_role": role_label(login_role),
                "login_active": bool(login_active),
                "password_changed": password_changed,
                "synced_bronco_certifications": synced_certifications,
            },
        )

        conn.commit()
        if account_fields_present:
            invalidate_sessions_for_card(
                card_id,
                except_token=(
                    token_from_request()
                    if account and account["card_id"] == card_id
                    else ""
                ),
            )
        include_sensitive = role == "admin"
        cards = card_rows(
            conn,
            include_accounts=include_sensitive,
            include_bronco_id=include_sensitive,
        )
        certifications = certification_rows(
            conn,
            include_bronco_id=include_sensitive,
        )
        people = people_rows(conn, include_bronco_id=include_sensitive)
        conn.close()

    return jsonify(
        {
            "ok": True,
            "cards": cards,
            "certifications": certifications,
            "people": people,
        }
    )


@app.post("/api/active-people/<card_id>/checkout")
def manual_checkout_person(card_id):
    role, error = require_access("staff")
    if error:
        return error

    account = account_for_token(token_from_request())
    data = request_json()
    try:
        card_id = validate_text(card_id, "card_id", MAX_ID_LENGTH, required=True)
        exited_at_value = validate_text(
            data.get("exited_at"),
            "exited_at",
            64,
            required=True,
        )
        exited_at = parse_iso(exited_at_value).replace(microsecond=0)
    except (ValueError, TypeError) as error:
        return jsonify({"ok": False, "error": f"Invalid checkout time: {error}"}), 400

    if exited_at > datetime.now(timezone.utc).replace(microsecond=0):
        return jsonify({"ok": False, "error": "Checkout time cannot be in the future"}), 400

    with db_lock:
        conn = db_connect()
        active_person = conn.execute(
            """
            SELECT card_id, entered_at, entry_door_id, entry_door_name
            FROM active_people
            WHERE card_id = ?
            """,
            (card_id,),
        ).fetchone()
        if not active_person:
            conn.close()
            return jsonify({"ok": False, "error": "Person is not currently inside"}), 404

        entered_at = parse_iso(active_person["entered_at"])
        if exited_at < entered_at:
            conn.close()
            return jsonify(
                {"ok": False, "error": "Checkout time cannot be before entry time"}
            ), 400

        card = card_row(conn, card_id)
        open_sessions = active_station_sessions_for_card(conn, card_id)
        conn.execute("DELETE FROM active_people WHERE card_id = ?", (card_id,))

        for session in open_sessions:
            conn.execute("DELETE FROM active_sessions WHERE id = ?", (session["id"],))
            log_event(
                conn,
                card_id,
                session["station_id"],
                session["station_name"],
                "station",
                "station_auto_out",
                duration_seconds=elapsed_seconds(session["started_at"], exited_at),
                details="Closed during manual checkout",
                event_time=exited_at.isoformat(),
            )

        exit_event = log_event(
            conn,
            card_id,
            active_person["entry_door_id"],
            active_person["entry_door_name"],
            "door",
            "manual_exit",
            duration_seconds=elapsed_seconds(active_person["entered_at"], exited_at),
            details="Manual checkout",
            event_time=exited_at.isoformat(),
        )
        add_audit_log(
            conn,
            account,
            "person_manually_checked_out",
            "person",
            card_id,
            {
                "name": card_display(card, card_id),
                "entered_at": active_person["entered_at"],
                "exited_at": exited_at.isoformat(),
                "closed_station_sessions": len(open_sessions),
            },
        )
        conn.commit()
        conn.close()

    return jsonify(
        {
            "ok": True,
            "event": exit_event,
            "closed_station_sessions": len(open_sessions),
        }
    )


@app.post("/api/warnings/clear")
def clear_warnings():
    role, error = require_access("staff")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        cursor = conn.execute(
            """
            INSERT OR IGNORE INTO warning_dismissals (
                swipe_event_id, dismissed_at, dismissed_by_card_id
            )
            SELECT id, ?, ?
            FROM swipe_events
            WHERE warning != ''
              AND action != 'station_auto_out'
            """,
            (now_iso(), account_field(account, "card_id")),
        )
        dismissed_count = max(0, cursor.rowcount)
        add_audit_log(
            conn,
            account,
            "warnings_cleared",
            "warning_feed",
            "live",
            {"dismissed_count": dismissed_count},
        )
        conn.commit()
        conn.close()

    return jsonify({"ok": True, "dismissed_count": dismissed_count})


@app.delete("/api/cards/<card_id>")
def delete_card(card_id):
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    try:
        card_id = validate_text(
            card_id,
            "card_id",
            MAX_ID_LENGTH,
            required=True,
        )
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400

    with db_lock:
        conn = db_connect()
        existing = card_row(conn, card_id)
        if not existing:
            conn.close()
            return jsonify({"ok": False, "error": "Card not found"}), 404

        if account and account.get("card_id") == card_id:
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "You cannot delete the card for the account you are currently signed in with",
                }
            ), 409

        related_counts = {
            "active_sessions": conn.execute(
                "SELECT COUNT(*) AS count FROM active_sessions WHERE card_id = ?",
                (card_id,),
            ).fetchone()["count"],
            "active_people": conn.execute(
                "SELECT COUNT(*) AS count FROM active_people WHERE card_id = ?",
                (card_id,),
            ).fetchone()["count"],
            "certifications": conn.execute(
                "SELECT COUNT(*) AS count FROM certifications WHERE card_id = ?",
                (card_id,),
            ).fetchone()["count"],
            "certify_permissions": conn.execute(
                "SELECT COUNT(*) AS count FROM certify_permissions WHERE card_id = ?",
                (card_id,),
            ).fetchone()["count"],
            "dashboard_login": conn.execute(
                "SELECT COUNT(*) AS count FROM user_accounts WHERE card_id = ?",
                (card_id,),
            ).fetchone()["count"],
        }

        conn.execute("DELETE FROM active_sessions WHERE card_id = ?", (card_id,))
        conn.execute("DELETE FROM active_people WHERE card_id = ?", (card_id,))
        conn.execute("DELETE FROM certifications WHERE card_id = ?", (card_id,))
        conn.execute("DELETE FROM certify_permissions WHERE card_id = ?", (card_id,))
        conn.execute("DELETE FROM user_accounts WHERE card_id = ?", (card_id,))
        conn.execute("DELETE FROM cards WHERE card_id = ?", (card_id,))
        conn.execute(
            "UPDATE canvas_sync_tasks SET card_id = '' WHERE card_id = ?",
            (card_id,),
        )
        latest_event_id = conn.execute(
            "SELECT COALESCE(MAX(id), 0) AS id FROM swipe_events WHERE card_id = ?",
            (card_id,),
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO pending_card_dismissals (
                card_id, ignored_through, ignored_through_event_id
            )
            VALUES (?, ?, ?)
            ON CONFLICT(card_id) DO UPDATE SET
                ignored_through = excluded.ignored_through,
                ignored_through_event_id = excluded.ignored_through_event_id
            """,
            (card_id, now_iso(), latest_event_id),
        )

        invalidate_sessions_for_card(card_id)

        for station_id, mode in list(cert_modes.items()):
            if mode.get("grantor_card_id") == card_id:
                cert_modes.pop(station_id, None)

        add_audit_log(
            conn,
            account,
            "card_deleted",
            "card",
            card_id,
            {
                "name": existing["name"],
                "email": existing["email"],
                "bronco_id": existing["bronco_id"],
                "designation": existing["designation"],
                "swipe_events_preserved": True,
                **related_counts,
            },
        )

        conn.commit()
        cards = card_rows(
            conn,
            include_accounts=True,
            include_bronco_id=True,
        )
        certifications = certification_rows(conn, include_bronco_id=True)
        people = people_rows(conn, include_bronco_id=True)
        certify_permissions = certify_permission_rows(conn)
        audit_log = audit_log_rows(conn)
        conn.close()

    return jsonify(
        {
            "ok": True,
            "cards": cards,
            "people": people,
            "certifications": certifications,
            "certify_permissions": certify_permissions,
            "audit_log": audit_log,
        }
    )


@app.post("/api/stations")
def save_station_rules():
    role, error = require_access("staff")
    if error:
        return error

    account = account_for_token(token_from_request())
    data = request_json()
    try:
        station_id = validate_text(
            data.get("station_id"),
            "station_id",
            MAX_ID_LENGTH,
            required=True,
        )
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400

    with db_lock:
        conn = db_connect()
        station = conn.execute(
            """
            SELECT
                id,
                kind,
                dashboard_visible,
                requires_certification,
                cert_override_active,
                cert_override_by,
                cert_override_updated_at,
                cert_override_expires_at
            FROM stations
            WHERE id = ?
            """,
            (station_id,),
        ).fetchone()

        if not station or station["kind"] != "station":
            conn.close()
            return jsonify({"ok": False, "error": "station_id must be a station"}), 400

        dashboard_visible = station["dashboard_visible"]
        requires_certification = station["requires_certification"]
        override_active = station["cert_override_active"]
        override_by = station["cert_override_by"] or ""
        override_updated_at = station["cert_override_updated_at"] or ""
        override_expires_at = station["cert_override_expires_at"] or ""

        if "dashboard_visible" in data:
            if role != "admin":
                conn.close()
                return jsonify(
                    {"ok": False, "error": "Admin login required to change dashboard visibility"}
                ), 403
            dashboard_visible = 1 if parse_bool(data.get("dashboard_visible")) else 0

        if "requires_certification" in data:
            requires_certification = (
                1 if parse_bool(data.get("requires_certification")) else 0
            )

        if "cert_override_active" in data:
            override_updated_at = now_iso()
            if parse_bool(data.get("cert_override_active")):
                if not requires_certification:
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Certification must be required before enabling an override",
                        }
                    ), 400
                try:
                    duration_minutes = int(
                        data.get("cert_override_duration_minutes", 30)
                    )
                except (TypeError, ValueError):
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Override duration must be a number of minutes",
                        }
                    ), 400

                if duration_minutes < 1 or duration_minutes > 1440:
                    conn.close()
                    return jsonify(
                        {
                            "ok": False,
                            "error": "Override duration must be between 1 and 1440 minutes",
                        }
                    ), 400
                expires_at = datetime.now(timezone.utc).replace(
                    microsecond=0
                ) + timedelta(minutes=duration_minutes)
                override_active = 1
                override_by = account["card_id"] if account else role
                override_expires_at = expires_at.isoformat()
            else:
                override_active = 0
                override_by = ""
                override_expires_at = ""

        if not requires_certification:
            override_active = 0
            override_by = ""
            override_expires_at = ""
            cert_modes.pop(station_id, None)

        conn.execute(
            """
            UPDATE stations
            SET
                dashboard_visible = ?,
                requires_certification = ?,
                cert_override_active = ?,
                cert_override_by = ?,
                cert_override_updated_at = ?,
                cert_override_expires_at = ?
            WHERE id = ?
            """,
            (
                dashboard_visible,
                requires_certification,
                override_active,
                override_by,
                override_updated_at,
                override_expires_at,
                station_id,
            ),
        )
        add_audit_log(
            conn,
            account,
            "station_rules_updated",
            "station",
            station_id,
            {
                "dashboard_visible": bool(dashboard_visible),
                "requires_certification": bool(requires_certification),
                "cert_override_active": bool(override_active),
                "cert_override_by": override_by,
                "cert_override_expires_at": override_expires_at,
            },
        )
        conn.commit()
        stations = station_option_rows(conn)
        conn.close()

    return jsonify({"ok": True, "stations": stations})


@app.post("/api/certify-permissions")
def save_certify_permissions():
    role, error = require_access("admin")
    if error:
        return error

    actor = account_for_token(token_from_request())
    data = request_json()
    try:
        card_id = validate_text(
            data.get("card_id"),
            "card_id",
            MAX_ID_LENGTH,
            required=True,
        )
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400

    raw_station_ids = data.get("station_ids", [])
    if not isinstance(raw_station_ids, list):
        return jsonify({"ok": False, "error": "station_ids must be a list"}), 400
    try:
        station_ids = [
            validate_text(item, "station_id", MAX_ID_LENGTH, required=True)
            for item in raw_station_ids
        ]
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400
    station_ids = list(dict.fromkeys(station_ids))

    with db_lock:
        conn = db_connect()
        account = conn.execute(
            """
            SELECT card_id, role
            FROM user_accounts
            WHERE card_id = ? AND active = 1
            """,
            (card_id,),
        ).fetchone()

        if not account or normalize_access_role(account["role"]) not in (
            "staff",
            "volunteer",
        ):
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "Choose an active Staff or Volunteer login",
                }
            ), 400

        valid_station_ids = {
            row["id"]
            for row in conn.execute(
                """
                SELECT id
                FROM stations
                WHERE kind = 'station'
                """
            ).fetchall()
        }
        if any(station_id not in valid_station_ids for station_id in station_ids):
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": "One or more station IDs are invalid",
                }
            ), 400

        conn.execute("DELETE FROM certify_permissions WHERE card_id = ?", (card_id,))
        timestamp = now_iso()
        for station_id in station_ids:
            conn.execute(
                """
                INSERT INTO certify_permissions (card_id, station_id, updated_at)
                VALUES (?, ?, ?)
                """,
                (card_id, station_id, timestamp),
            )

        add_audit_log(
            conn,
            actor,
            "certify_permissions_updated",
            "card",
            card_id,
            {"station_ids": station_ids},
        )
        conn.commit()
        rows = certify_permission_rows(conn)
        conn.close()

    return jsonify({"ok": True, "certify_permissions": rows})


@app.post("/api/certifications")
def save_certification():
    role, error = require_access("staff", "volunteer")
    if error:
        return error

    account = account_for_token(token_from_request())
    data = request_json()
    try:
        card_id = validate_text(
            data.get("card_id"),
            "card_id",
            MAX_ID_LENGTH,
            required=True,
        )
        station_id = validate_text(
            data.get("station_id"),
            "station_id",
            MAX_ID_LENGTH,
            required=True,
        )
        notes = validate_text(data.get("notes"), "notes", MAX_NOTES_LENGTH)
    except ValueError as error:
        return jsonify({"ok": False, "error": str(error)}), 400
    active = 1 if parse_bool(data.get("active", True)) else 0

    with db_lock:
        conn = db_connect()
        card = card_row(conn, card_id)
        station = conn.execute(
            """
            SELECT id, kind
            FROM stations
            WHERE id = ?
            """,
            (station_id,),
        ).fetchone()

        if not card:
            conn.close()
            return jsonify({"ok": False, "error": "card_id is not in card database"}), 404

        if not station or station["kind"] != "station":
            conn.close()
            return jsonify({"ok": False, "error": "station_id must be a station"}), 400

        if not account_can_certify_station(conn, account, station_id):
            conn.close()
            return jsonify(
                {
                    "ok": False,
                    "error": (
                        "You are not authorized to manage certifications "
                        "for this station"
                    ),
                }
            ), 403

        timestamp = now_iso()
        granted_by = account["card_id"] if account else ""
        conn.execute(
            """
            INSERT INTO certifications (
                card_id,
                station_id,
                active,
                notes,
                updated_at,
                granted_via,
                granted_by,
                granted_at
            )
            VALUES (?, ?, ?, ?, ?, 'dashboard', ?, ?)
            ON CONFLICT(card_id, station_id) DO UPDATE SET
                active = excluded.active,
                notes = excluded.notes,
                updated_at = excluded.updated_at,
                granted_via = CASE
                    WHEN excluded.active = 1
                    THEN excluded.granted_via
                    ELSE certifications.granted_via
                END,
                granted_by = CASE
                    WHEN excluded.active = 1
                    THEN excluded.granted_by
                    ELSE certifications.granted_by
                END,
                granted_at = CASE
                    WHEN excluded.active = 1
                    THEN excluded.granted_at
                    ELSE certifications.granted_at
                END
            """,
            (card_id, station_id, active, notes, timestamp, granted_by, timestamp),
        )
        if card["bronco_id"]:
            conn.execute(
                """
                INSERT INTO bronco_certifications (
                    bronco_id, station_id, active, notes, updated_at
                )
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(bronco_id, station_id) DO UPDATE SET
                    active = excluded.active,
                    notes = excluded.notes,
                    updated_at = excluded.updated_at
                """,
                (card["bronco_id"], station_id, active, notes, timestamp),
            )
            queue_canvas_sync_task(conn, card, station_id, bool(active), account)
        add_audit_log(
            conn,
            account,
            "certification_granted" if active else "certification_revoked",
            "certification",
            f"{card_id}:{station_id}",
            {
                "card_id": card_id,
                "station_id": station_id,
                "active": bool(active),
                "notes": notes,
            },
        )
        conn.commit()
        certifications = certification_rows(
            conn,
            include_bronco_id=(role == "admin"),
        )
        conn.close()

    return jsonify({"ok": True, "certifications": certifications})


@app.get("/status")
def status():
    account = account_for_token(token_from_request())
    role = normalize_access_role(account["role"]) if account else ""
    include_people = role in ("admin", "staff")
    with db_lock:
        conn = db_connect()
        rows = station_status_rows(conn)
        conn.close()
    if not include_people:
        rows = [{**row, "active_cards": ""} for row in rows]
    return jsonify(rows)


def card_rows(conn, include_accounts=False, include_bronco_id=False):
    if include_accounts:
        rows = conn.execute(
            """
            SELECT
                cards.card_id,
                cards.bronco_id,
                cards.name,
                cards.email,
                cards.designation,
                cards.active,
                cards.notes,
                cards.updated_at,
                user_accounts.username AS login_username,
                user_accounts.role AS login_role,
                user_accounts.active AS login_active,
                user_accounts.card_id AS login_card_id
            FROM cards
            LEFT JOIN user_accounts
                ON user_accounts.card_id = cards.card_id
            ORDER BY
                CASE WHEN cards.name = '' THEN 1 ELSE 0 END,
                cards.name,
                cards.card_id
            """
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT card_id, bronco_id, name, email, designation, active, notes, updated_at
            FROM cards
            ORDER BY
                CASE WHEN name = '' THEN 1 ELSE 0 END,
                name,
                card_id
            """
        ).fetchall()

    result = []
    for row in rows:
        card = {
            "card_id": row["card_id"],
            "person_ref": person_ref_for_bronco_id(row["bronco_id"]),
            "name": row["name"],
            "email": row["email"],
            "designation": row["designation"],
            "active": bool(row["active"]),
            "notes": row["notes"],
            "updated_at": row["updated_at"],
            "display_name": card_display(row, row["card_id"]),
        }

        if include_bronco_id:
            card["bronco_id"] = row["bronco_id"]

        if include_accounts:
            card["has_login"] = bool(row["login_card_id"])
            card["login_username"] = row["login_username"] or ""
            card["login_role"] = row["login_role"] or ""
            card["login_active"] = (
                bool(row["login_active"])
                if row["login_card_id"]
                else False
            )

        result.append(card)

    return result

def audit_log_rows(conn, limit=80):
    rows = conn.execute(
        """
        SELECT
            id,
            timestamp,
            actor_card_id,
            actor_name,
            actor_role,
            action,
            target_type,
            target_id,
            details
        FROM audit_log
        ORDER BY id DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()

    result = []
    for row in rows:
        try:
            details = json.loads(row["details"]) if row["details"] else {}
        except json.JSONDecodeError:
            details = {"details": row["details"]}

        result.append(
            {
                "id": row["id"],
                "timestamp": row["timestamp"],
                "actor_card_id": row["actor_card_id"],
                "actor_name": row["actor_name"],
                "actor_role": row["actor_role"],
                "action": row["action"],
                "target_type": row["target_type"],
                "target_id": row["target_id"],
                "details": details,
            }
        )

    return result


def station_option_rows(conn):
    rows = conn.execute(
        """
        SELECT
            id AS station_id,
            name AS station_name,
            dashboard_visible,
            requires_certification,
            cert_override_active,
            cert_override_by,
            cert_override_updated_at,
            cert_override_expires_at
        FROM stations
        WHERE kind = 'station'
        ORDER BY name
        """
    ).fetchall()
    return [
        {
            "station_id": row["station_id"],
            "station_name": row["station_name"],
            "dashboard_visible": bool(row["dashboard_visible"]),
            "requires_certification": bool(row["requires_certification"]),
            "cert_override_active": override_is_effective(row),
            "cert_override_by": row["cert_override_by"] or "",
            "cert_override_updated_at": row["cert_override_updated_at"] or "",
            "cert_override_expires_at": row["cert_override_expires_at"] or "",
        }
        for row in rows
    ]


def reader_option_rows(conn):
    rows = conn.execute(
        """
        SELECT id AS station_id, name AS station_name, kind AS station_kind
        FROM stations
        ORDER BY
            CASE WHEN kind = 'door' THEN 0 ELSE 1 END,
            name
        """
    ).fetchall()
    return [dict(row) for row in rows]


def certification_rows(conn, include_bronco_id=False):
    rows = conn.execute(
        """
        SELECT
            certifications.card_id,
            cards.bronco_id,
            cards.name,
            cards.email,
            cards.designation,
            certifications.station_id,
            stations.name AS station_name,
            certifications.active,
            certifications.notes,
            certifications.updated_at,
            certifications.granted_via,
            certifications.granted_by,
            certifications.granted_at
        FROM certifications
        JOIN cards ON cards.card_id = certifications.card_id
        JOIN stations ON stations.id = certifications.station_id
        ORDER BY cards.name, certifications.card_id, stations.name
        """
    ).fetchall()

    result = []
    for row in rows:
        certification = {
            "card_id": row["card_id"],
            "person_ref": person_ref_for_bronco_id(row["bronco_id"]),
            "name": row["name"],
            "email": row["email"],
            "designation": row["designation"],
            "display_name": card_display(row, row["card_id"]),
            "station_id": row["station_id"],
            "station_name": row["station_name"],
            "active": bool(row["active"]),
            "notes": row["notes"],
            "updated_at": row["updated_at"],
            "granted_via": row["granted_via"],
            "granted_by": row["granted_by"],
            "granted_at": row["granted_at"],
        }
        if include_bronco_id:
            certification["bronco_id"] = row["bronco_id"]
        result.append(certification)

    return result

def active_people_rows(conn):
    current_time = datetime.now(timezone.utc)
    rows = conn.execute(
        """
        SELECT
            active_people.card_id,
            cards.name,
            cards.email,
            cards.designation,
            active_people.entered_at,
            active_people.entry_door_id,
            active_people.entry_door_name
        FROM active_people
        LEFT JOIN cards ON cards.card_id = active_people.card_id
        ORDER BY entered_at
        """
    ).fetchall()

    result = []
    for row in rows:
        result.append(
            {
                "card_id": row["card_id"],
                "name": row["name"] or "",
                "email": row["email"] or "",
                "designation": row["designation"] or "",
                "display_name": card_display(row, row["card_id"]),
                "entered_at": row["entered_at"],
                "entry_door_id": row["entry_door_id"],
                "entry_door_name": row["entry_door_name"],
                "elapsed_seconds": elapsed_seconds(row["entered_at"], current_time),
            }
        )

    return result


def station_status_rows(conn):
    rows = conn.execute(
        """
        SELECT
            stations.id AS station_id,
            stations.name AS station_name,
            COUNT(active_sessions.id) AS active_sessions,
            GROUP_CONCAT(
                COALESCE(NULLIF(cards.name, ''), active_sessions.card_id),
                '; '
            ) AS active_cards
        FROM stations
        LEFT JOIN active_sessions
            ON active_sessions.station_id = stations.id
        LEFT JOIN cards
            ON cards.card_id = active_sessions.card_id
        WHERE stations.kind = 'station'
          AND stations.dashboard_visible = 1
        GROUP BY stations.id, stations.name
        ORDER BY
            COUNT(active_sessions.id) DESC,
            CASE stations.id
                WHEN '3d-printing' THEN 1
                WHEN 'soldering' THEN 2
                WHEN 'vinyl' THEN 3
                WHEN 'laser-cutting' THEN 4
                WHEN 'sewing' THEN 5
                WHEN 'embroidery' THEN 6
                ELSE 99
            END,
            stations.name
        """
    ).fetchall()

    return [
        {
            "station_id": row["station_id"],
            "station_name": row["station_name"],
            "active_sessions": int(row["active_sessions"]),
            "active_cards": row["active_cards"] or "",
        }
        for row in rows
    ]


def active_session_rows(conn):
    current_time = datetime.now(timezone.utc)
    rows = conn.execute(
        """
        SELECT
            active_sessions.card_id,
            cards.name,
            cards.email,
            cards.designation,
            active_sessions.station_id,
            active_sessions.station_name,
            active_sessions.started_at
        FROM active_sessions
        LEFT JOIN cards ON cards.card_id = active_sessions.card_id
        ORDER BY station_id, started_at
        """
    ).fetchall()

    result = []
    for row in rows:
        result.append(
            {
                "card_id": row["card_id"],
                "name": row["name"] or "",
                "email": row["email"] or "",
                "designation": row["designation"] or "",
                "display_name": card_display(row, row["card_id"]),
                "station_id": row["station_id"],
                "station_name": row["station_name"],
                "started_at": row["started_at"],
                "elapsed_seconds": elapsed_seconds(row["started_at"], current_time),
            }
        )

    return result


def recent_swipe_rows(conn, limit, query="", station_id="", warning_only=False):
    clauses = []
    params = []
    if query:
        like = f"%{query.lower()}%"
        clauses.append("(lower(cards.name) LIKE ? OR lower(cards.email) LIKE ? OR lower(swipe_events.card_id) LIKE ? OR lower(swipe_events.action) LIKE ?)")
        params.extend([like, like, like, like])
    if station_id:
        clauses.append("swipe_events.station_id = ?")
        params.append(station_id)
    if warning_only:
        clauses.append("swipe_events.warning != ''")
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = conn.execute(
        f"""
        SELECT
            swipe_events.card_id,
            cards.name,
            cards.email,
            cards.designation,
            swipe_events.station_id,
            swipe_events.station_name,
            swipe_events.station_kind,
            swipe_events.timestamp,
            swipe_events.action,
            swipe_events.allowed,
            swipe_events.duration_seconds,
            swipe_events.active_users,
            swipe_events.warning,
            swipe_events.details
        FROM swipe_events
        LEFT JOIN cards ON cards.card_id = swipe_events.card_id
        {where}
        ORDER BY id DESC
        LIMIT ?
        """,
        (*params, limit),
    ).fetchall()

    return [
        {
            **dict(row),
            "name": row["name"] or "",
            "email": row["email"] or "",
            "designation": row["designation"] or "",
            "display_name": card_display(row, row["card_id"]),
            "allowed": bool(row["allowed"]),
        }
        for row in rows
    ]


def warning_rows(conn, limit):
    rows = conn.execute(
        """
        SELECT
            swipe_events.card_id,
            cards.name,
            cards.email,
            cards.designation,
            swipe_events.station_id,
            swipe_events.station_name,
            swipe_events.timestamp,
            swipe_events.action,
            swipe_events.warning,
            swipe_events.details
        FROM swipe_events
        LEFT JOIN cards ON cards.card_id = swipe_events.card_id
        LEFT JOIN warning_dismissals
            ON warning_dismissals.swipe_event_id = swipe_events.id
        WHERE warning != ''
          AND swipe_events.action != 'station_auto_out'
          AND warning_dismissals.swipe_event_id IS NULL
        ORDER BY id DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()

    return [
        {
            **dict(row),
            "name": row["name"] or "",
            "email": row["email"] or "",
            "designation": row["designation"] or "",
            "display_name": card_display(row, row["card_id"]),
        }
        for row in rows
    ]


def analytics_snapshot(conn, days=30, current_time=None):
    current_time = current_time or datetime.now(timezone.utc)
    cutoff = current_time - timedelta(days=days) if days else None
    cutoff_iso = cutoff.replace(microsecond=0).isoformat() if cutoff else ""

    station_rows = conn.execute(
        "SELECT id, name FROM stations WHERE kind = 'station' ORDER BY name"
    ).fetchall()
    station_usage = {
        row["id"]: {
            "station_id": row["id"],
            "station_name": row["name"],
            "sessions": 0,
            "total_seconds": 0,
            "unique_card_ids": set(),
            "active_now": 0,
        }
        for row in station_rows
    }

    event_query = """
        SELECT station_id, station_name, card_id, duration_seconds
        FROM swipe_events
        WHERE station_kind = 'station'
          AND allowed = 1
          AND action IN ('station_out', 'station_auto_out')
    """
    event_params = []
    if cutoff_iso:
        event_query += " AND timestamp >= ?"
        event_params.append(cutoff_iso)

    for row in conn.execute(event_query, event_params).fetchall():
        item = station_usage.setdefault(
            row["station_id"],
            {
                "station_id": row["station_id"],
                "station_name": row["station_name"],
                "sessions": 0,
                "total_seconds": 0,
                "unique_card_ids": set(),
                "active_now": 0,
            },
        )
        item["sessions"] += 1
        item["total_seconds"] += max(0, int(row["duration_seconds"] or 0))
        item["unique_card_ids"].add(row["card_id"])

    for row in conn.execute(
        "SELECT station_id, COUNT(*) AS active_now FROM active_sessions GROUP BY station_id"
    ).fetchall():
        if row["station_id"] in station_usage:
            station_usage[row["station_id"]]["active_now"] = row["active_now"]

    usage_rows = []
    all_unique_cards = set()
    for item in station_usage.values():
        all_unique_cards.update(item["unique_card_ids"])
        sessions = item["sessions"]
        usage_rows.append(
            {
                "station_id": item["station_id"],
                "station_name": item["station_name"],
                "sessions": sessions,
                "total_seconds": item["total_seconds"],
                "average_seconds": int(item["total_seconds"] / sessions) if sessions else 0,
                "unique_users": len(item["unique_card_ids"]),
                "active_now": item["active_now"],
            }
        )
    usage_rows.sort(
        key=lambda item: (
            -item["total_seconds"],
            -item["sessions"],
            item["station_name"].lower(),
        )
    )

    station_hour_counts = [0] * 24
    arrival_hour_counts = [0] * 24
    station_weekday_counts = [0] * 7
    arrival_weekday_counts = [0] * 7
    visitor_entries = 0
    unique_visitors = set()
    activity_query = """
        SELECT timestamp, card_id, station_kind, action
        FROM swipe_events
        WHERE allowed = 1
          AND (
            (station_kind = 'station' AND action = 'station_in')
            OR (station_kind = 'door' AND action = 'enter')
          )
    """
    activity_params = []
    if cutoff_iso:
        activity_query += " AND timestamp >= ?"
        activity_params.append(cutoff_iso)
    for row in conn.execute(activity_query, activity_params).fetchall():
        try:
            local_time = local_space_time(parse_iso(row["timestamp"]))
        except (TypeError, ValueError, OverflowError):
            continue
        if row["station_kind"] == "station":
            station_hour_counts[local_time.hour] += 1
            station_weekday_counts[local_time.weekday()] += 1
        else:
            arrival_hour_counts[local_time.hour] += 1
            arrival_weekday_counts[local_time.weekday()] += 1
            visitor_entries += 1
            unique_visitors.add(row["card_id"])

    weekday_names = (
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    )
    hourly_activity = [
        {
            "hour": hour,
            "label": display_hour(hour),
            "station_starts": station_hour_counts[hour],
            "visitor_entries": arrival_hour_counts[hour],
        }
        for hour in range(24)
    ]
    weekday_activity = [
        {
            "weekday": weekday,
            "label": weekday_names[weekday],
            "station_starts": station_weekday_counts[weekday],
            "visitor_entries": arrival_weekday_counts[weekday],
        }
        for weekday in range(7)
    ]
    busiest_station_hour = (
        max(range(24), key=station_hour_counts.__getitem__)
        if any(station_hour_counts)
        else None
    )
    busiest_arrival_hour = (
        max(range(24), key=arrival_hour_counts.__getitem__)
        if any(arrival_hour_counts)
        else None
    )
    busiest_weekday = (
        max(range(7), key=station_weekday_counts.__getitem__)
        if any(station_weekday_counts)
        else None
    )

    total_sessions = sum(item["sessions"] for item in usage_rows)
    total_usage_seconds = sum(item["total_seconds"] for item in usage_rows)
    top_station = next((item for item in usage_rows if item["sessions"]), None)
    return {
        "generated_at": current_time.replace(microsecond=0).isoformat(),
        "days": days,
        "period_label": "All time" if not days else f"Last {days} days",
        "timezone": SPACE_TIMEZONE_NAME,
        "opening_hours": opening_hours_label(),
        "space_open_now": space_is_open(current_time),
        "summary": {
            "completed_sessions": total_sessions,
            "total_usage_seconds": total_usage_seconds,
            "unique_users": len(all_unique_cards),
            "visitor_entries": visitor_entries,
            "unique_visitors": len(unique_visitors),
            "busiest_station": top_station["station_name"] if top_station else "--",
            "busiest_hour": (
                display_hour(busiest_station_hour)
                if busiest_station_hour is not None
                else "--"
            ),
            "busiest_arrival_hour": (
                display_hour(busiest_arrival_hour)
                if busiest_arrival_hour is not None
                else "--"
            ),
            "busiest_weekday": (
                weekday_names[busiest_weekday]
                if busiest_weekday is not None
                else "--"
            ),
        },
        "station_usage": usage_rows,
        "hourly_activity": hourly_activity,
        "weekday_activity": weekday_activity,
    }


def workbook_safe_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (dict, list)):
        value = json.dumps(value, sort_keys=True, separators=(",", ":"))
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + value
    return value


def add_workbook_sheet(workbook, title, headers, rows):
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")

    widths = [len(str(header)) for header in headers]
    for raw_row in rows:
        row = dict(raw_row)
        values = [workbook_safe_cell(row.get(header, "")) for header in headers]
        worksheet.append(values)
        for index, value in enumerate(values):
            widths[index] = min(60, max(widths[index], len(str(value))))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(10, width + 2)
    return worksheet


def build_master_workbook(conn):
    workbook = Workbook()
    workbook.remove(workbook.active)
    people = conn.execute(
        """
        SELECT people.bronco_id, people.name, people.email, people.designation,
               people.active, cards.card_id AS assigned_card_id,
               people.notes, people.updated_at
        FROM people
        LEFT JOIN cards ON lower(cards.bronco_id) = lower(people.bronco_id)
        ORDER BY people.name, people.bronco_id
        """
    ).fetchall()
    add_workbook_sheet(
        workbook, "People",
        ["bronco_id", "name", "email", "designation", "active", "assigned_card_id", "notes", "updated_at"],
        people,
    )

    certifications = conn.execute(
        """
        SELECT bronco_certifications.bronco_id, people.name, people.email,
               cards.card_id, bronco_certifications.station_id,
               stations.name AS station_name, bronco_certifications.active,
               bronco_certifications.updated_at, certifications.granted_via,
               certifications.granted_by, certifications.granted_at,
               bronco_certifications.notes
        FROM bronco_certifications
        JOIN people ON people.bronco_id = bronco_certifications.bronco_id
        JOIN stations ON stations.id = bronco_certifications.station_id
        LEFT JOIN cards ON lower(cards.bronco_id) = lower(people.bronco_id)
        LEFT JOIN certifications
          ON certifications.card_id = cards.card_id
         AND certifications.station_id = bronco_certifications.station_id
        ORDER BY people.name, stations.name
        """
    ).fetchall()
    add_workbook_sheet(
        workbook, "Certifications",
        ["bronco_id", "name", "email", "card_id", "station_id", "station_name", "active", "updated_at", "granted_via", "granted_by", "granted_at", "notes"],
        certifications,
    )

    cards = conn.execute(
        """
        SELECT cards.card_id, cards.bronco_id, cards.name, cards.email,
               cards.designation, cards.active,
               user_accounts.username AS login_username,
               user_accounts.role AS login_role,
               user_accounts.active AS login_active,
               cards.notes, cards.updated_at
        FROM cards
        LEFT JOIN user_accounts ON user_accounts.card_id = cards.card_id
        ORDER BY cards.name, cards.card_id
        """
    ).fetchall()
    add_workbook_sheet(
        workbook, "Cards",
        ["card_id", "bronco_id", "name", "email", "designation", "active", "login_username", "login_role", "login_active", "notes", "updated_at"],
        cards,
    )

    add_workbook_sheet(
        workbook, "Pending Canvas",
        ["bronco_id", "name", "email", "card_id", "station_id", "station_name", "desired_active", "created_at", "created_by_card_id", "created_by_name"],
        canvas_sync_task_rows(conn, include_bronco_id=True),
    )

    swipe_rows = conn.execute(
        """
        SELECT swipe_events.timestamp, swipe_events.card_id, cards.bronco_id,
               cards.name, cards.email, cards.designation,
               swipe_events.station_id, swipe_events.station_name,
               swipe_events.station_kind, swipe_events.action,
               swipe_events.allowed, swipe_events.duration_seconds,
               swipe_events.active_users, swipe_events.warning,
               swipe_events.details, swipe_events.event_id
        FROM swipe_events
        LEFT JOIN cards ON cards.card_id = swipe_events.card_id
        ORDER BY swipe_events.id
        """
    ).fetchall()
    add_workbook_sheet(
        workbook, "Swipe Log",
        ["timestamp", "card_id", "bronco_id", "name", "email", "designation", "station_id", "station_name", "station_kind", "action", "allowed", "duration_seconds", "active_users", "warning", "details", "event_id"],
        swipe_rows,
    )

    audit_rows = conn.execute(
        """
        SELECT timestamp, actor_card_id, actor_name, actor_role, action,
               target_type, target_id, details
        FROM audit_log ORDER BY id
        """
    ).fetchall()
    add_workbook_sheet(
        workbook, "Audit Log",
        ["timestamp", "actor_card_id", "actor_name", "actor_role", "action", "target_type", "target_id", "details"],
        audit_rows,
    )

    analytics = analytics_snapshot(conn, days=30)
    add_workbook_sheet(
        workbook, "Station Analytics",
        ["station_id", "station_name", "sessions", "total_seconds", "average_seconds", "unique_users", "active_now"],
        analytics["station_usage"],
    )
    add_workbook_sheet(
        workbook, "Hourly Activity",
        ["hour", "label", "station_starts", "visitor_entries"],
        analytics["hourly_activity"],
    )
    add_workbook_sheet(
        workbook, "Weekday Activity",
        ["weekday", "label", "station_starts", "visitor_entries"],
        analytics["weekday_activity"],
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def csv_reply(filename, headers, rows):
    def safe_cell(value):
        if not isinstance(value, str):
            return value
        if value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
            return "'" + value
        return value

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(
        {
            key: safe_cell(value)
            for key, value in dict(row).items()
        }
        for row in rows
    )

    return Response(
        "\ufeff" + output.getvalue(),
        content_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/analytics")
def analytics_api():
    role, error = require_access("admin")
    if error:
        return error

    try:
        days = int(request.args.get("days", "30"))
    except ValueError:
        return jsonify({"ok": False, "error": "days must be 0, 7, 30, or 90"}), 400
    if days not in (0, 7, 30, 90):
        return jsonify({"ok": False, "error": "days must be 0, 7, 30, or 90"}), 400

    with db_lock:
        conn = db_connect()
        analytics = analytics_snapshot(conn, days=days)
        conn.close()
    return jsonify({"ok": True, "analytics": analytics})


@app.get("/master-export.xlsx")
def master_workbook_export():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        add_audit_log(
            conn,
            account,
            "master_workbook_exported",
            "export",
            "master-export.xlsx",
        )
        conn.commit()
        workbook = build_master_workbook(conn)
        conn.close()

    filename = f"station-master-{local_space_time().date().isoformat()}.xlsx"
    return Response(
        workbook.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/swipes.csv")
def swipes_csv():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        add_audit_log(conn, account, "swipe_log_exported", "export", "swipes.csv")
        conn.commit()
        rows = conn.execute(
            """
            SELECT
                swipe_events.card_id,
                cards.name,
                cards.email,
                cards.designation,
                swipe_events.station_id,
                swipe_events.station_name,
                swipe_events.station_kind,
                swipe_events.timestamp,
                swipe_events.action,
                swipe_events.allowed,
                swipe_events.duration_seconds,
                swipe_events.active_users,
                swipe_events.warning,
                swipe_events.details,
                swipe_events.event_id
            FROM swipe_events
            LEFT JOIN cards ON cards.card_id = swipe_events.card_id
            ORDER BY id
            """
        ).fetchall()
        conn.close()

    headers = [
        "card_id",
        "name",
        "email",
        "designation",
        "station_id",
        "station_name",
        "station_kind",
        "timestamp",
        "action",
        "allowed",
        "duration_seconds",
        "active_users",
        "warning",
        "details",
        "event_id",
    ]
    return csv_reply("swipes.csv", headers, [dict(row) for row in rows])


@app.get("/cards.csv")
def cards_csv():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        add_audit_log(conn, account, "card_database_exported", "export", "cards.csv")
        conn.commit()
        rows = card_rows(
            conn,
            include_accounts=True,
            include_bronco_id=True,
        )
        conn.close()

    headers = [
        "card_id",
        "name",
        "email",
        "designation",
        "active",
        "notes",
        "updated_at",
    ]
    headers[1:1] = ["bronco_id"]
    headers.extend(["login_username", "login_role", "login_active"])

    return csv_reply("cards.csv", headers, rows)


@app.get("/audit.csv")
def audit_csv():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        add_audit_log(conn, account, "audit_log_exported", "export", "audit.csv")
        conn.commit()
        rows = conn.execute(
            """
            SELECT
                timestamp,
                actor_card_id,
                actor_name,
                actor_role,
                action,
                target_type,
                target_id,
                details
            FROM audit_log
            ORDER BY id
            """
        ).fetchall()
        conn.close()

    headers = [
        "timestamp",
        "actor_card_id",
        "actor_name",
        "actor_role",
        "action",
        "target_type",
        "target_id",
        "details",
    ]
    return csv_reply("audit.csv", headers, [dict(row) for row in rows])


@app.get("/active.csv")
def active_csv():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        add_audit_log(conn, account, "active_people_exported", "export", "active.csv")
        conn.commit()
        rows = active_people_rows(conn)
        conn.close()

    headers = [
        "card_id",
        "name",
        "email",
        "designation",
        "display_name",
        "entered_at",
        "entry_door_id",
        "entry_door_name",
        "elapsed_seconds",
    ]
    return csv_reply("active.csv", headers, rows)


@app.get("/station_status.csv")
def station_status_csv():
    role, error = require_access("admin")
    if error:
        return error

    account = account_for_token(token_from_request())
    with db_lock:
        conn = db_connect()
        rows = station_status_rows(conn)
        conn.close()

    headers = ["station_id", "station_name", "active_sessions", "active_cards"]
    return csv_reply("station_status.csv", headers, rows)


init_db()
seed_stations()
bootstrap_admin()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
